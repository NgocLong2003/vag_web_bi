"""
Báo cáo Kinh Doanh — Blueprint (DuckDB version)
API prefix: /reports/bao-cao-kinh-doanh/api/...
"""
from flask import Blueprint, request, send_file, current_app
from api_logger import api_response, set_api_result
from datetime import datetime, date
import logging

logger = logging.getLogger(__name__)

bp = Blueprint('bckd', __name__,
               url_prefix='/reports/bao-cao-kinh-doanh',
               template_folder='templates')

from query_loader import load_sql


def get_store():
    return current_app.config['DUCKDB_STORE']


# ─────────────────────────────────────────
#  API: Hierarchy
# ─────────────────────────────────────────
@bp.route('/api/hierarchy')
def api_hierarchy():
    try:
        data = get_store().query(load_sql('HIERARCHY_CTE_DUCK'))
        return api_response(ok=True, data=data, count=len(data))
    except Exception as e:
        logger.error(f"[hierarchy] {e}")
        return api_response(ok=False, error=str(e))


# ─────────────────────────────────────────
#  API: Khách hàng
# ─────────────────────────────────────────
@bp.route('/api/khachhang')
def api_khachhang():
    try:
        data = get_store().query(load_sql('KHACHHANG_DUCK'))
        return api_response(ok=True, data=data, count=len(data))
    except Exception as e:
        logger.error(f"[khachhang] {e}")
        return api_response(ok=False, error=str(e))


# ─────────────────────────────────────────
#  API: Công nợ
# ─────────────────────────────────────────
@bp.route('/api/congno', methods=['POST'])
def api_congno():
    body = request.get_json(force=True)
    ngay_cut = body.get('ngay_cut', '')
    ma_bp = body.get('ma_bp', '')
    ds_nvkd = body.get('ds_nvkd', '')
    ds_kh = body.get('ds_kh', '')

    if not ngay_cut:
        return api_response(ok=False, error='Thiếu ngay_cut', status_code=400)
    try:
        start_y = datetime.strptime(ngay_cut, '%Y-%m-%d').year
    except ValueError:
        return api_response(ok=False, error='ngay_cut không hợp lệ', status_code=400)

    try:
        data = get_store().query(
            load_sql('CONGNO_SQL_DUCK'),
            [ngay_cut, start_y, ma_bp or '', ds_nvkd or '', ds_kh or '']
        )
        for d in data:
            for k in ('so_du_ban_dau', 'tong_phatsinh', 'du_no_ck'):
                if d.get(k) is not None:
                    d[k] = float(d[k])
        return api_response(ok=True, data=data, count=len(data),
                            meta={'ngay_cut': ngay_cut, 'ma_bp': ma_bp})
    except Exception as e:
        logger.error(f"[congno] {e}")
        return api_response(ok=False, error=str(e))


# ─────────────────────────────────────────
#  API: Doanh số
# ─────────────────────────────────────────
@bp.route('/api/doanhso', methods=['POST'])
def api_doanhso():
    body = request.get_json(force=True)
    ngay_a = body.get('ngay_a', '')
    ngay_b = body.get('ngay_b', '')
    ma_bp = body.get('ma_bp', '')
    ds_nvkd = body.get('ds_nvkd', '')
    ds_kh = body.get('ds_kh', '')

    if not ngay_a or not ngay_b:
        return api_response(ok=False, error='Thiếu ngay_a hoặc ngay_b', status_code=400)

    try:
        data = get_store().query(
            load_sql('DOANHSO_SQL_DUCK'),
            [ngay_a, ngay_b, ma_bp or '', ds_nvkd or '', ds_kh or '']
        )
        for d in data:
            for k in ('tong_so_luong', 'tong_tien_nt2', 'tong_tien_ck_nt', 'tong_thue_gtgt_nt', 'tong_doanhso'):
                if d.get(k) is not None:
                    d[k] = float(d[k])
        return api_response(ok=True, data=data, count=len(data),
                            meta={'ngay_a': ngay_a, 'ngay_b': ngay_b, 'ma_bp': ma_bp})
    except Exception as e:
        logger.error(f"[doanhso] {e}")
        return api_response(ok=False, error=str(e))


# ─────────────────────────────────────────
#  API: Doanh thu
# ─────────────────────────────────────────
@bp.route('/api/doanhthu', methods=['POST'])
def api_doanhthu():
    body = request.get_json(force=True)
    ngay_a = body.get('ngay_a', '')
    ngay_b = body.get('ngay_b', '')
    ma_bp = body.get('ma_bp', '')
    ds_nvkd = body.get('ds_nvkd', '')
    ds_kh = body.get('ds_kh', '')

    if not ngay_a or not ngay_b:
        return api_response(ok=False, error='Thiếu ngay_a hoặc ngay_b', status_code=400)

    try:
        data = get_store().query(
            load_sql('DOANHTHU_BCKD_DUCK'),
            [ngay_a, ngay_b, ma_bp or '', ds_nvkd or '', ds_kh or '']
        )
        for d in data:
            if d.get('doanhthu') is not None:
                d['doanhthu'] = float(d['doanhthu'])
        return api_response(ok=True, data=data, count=len(data),
                            meta={'ngay_a': ngay_a, 'ngay_b': ngay_b, 'ma_bp': ma_bp})
    except Exception as e:
        logger.error(f"[doanhthu] {e}")
        return api_response(ok=False, error=str(e))


# ─────────────────────────────────────────
#  API: Export Excel
# ─────────────────────────────────────────
@bp.route('/api/export_excel', methods=['POST'])
def api_export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    body = request.get_json(force=True)
    rows = body.get('rows', [])
    col_headers = body.get('col_headers', [])

    if not rows:
        return api_response(ok=False, error='Không có dữ liệu', status_code=400)

    max_nv_depth = 0
    for r in rows:
        if r.get('type') == 'nv':
            max_nv_depth = max(max_nv_depth, r.get('depth', 0))
    nv_cols = max_nv_depth + 1
    kh_col = nv_cols + 1
    data_start = nv_cols + 2

    wb = Workbook()
    ws = wb.active
    ws.title = 'Báo cáo'
    FONT_NAME = 'Arial'
    BLACK = '000000'
    thin_border = Border(bottom=Side(style='thin', color='D5D9E4'), right=Side(style='thin', color='ECEEF3'))
    header_border = Border(bottom=Side(style='medium', color='A0AAC0'))
    total_border = Border(top=Side(style='medium', color='8090B0'), bottom=Side(style='medium', color='8090B0'))
    nv_bg_colors = ['B8C6F0', 'CADAF6', 'DAEAFC', 'E8F0FD', 'F0F5FE', 'F7FAFE']

    def nv_bg(depth): return nv_bg_colors[min(depth, len(nv_bg_colors) - 1)]
    def nv_font_sz(depth): return 11 if depth == 0 else 10.5 if depth == 1 else 10

    cur_row = 1
    ws.cell(row=1, column=1, value='NHÂN VIÊN KINH DOANH')
    ws.cell(row=1, column=1).font = Font(name=FONT_NAME, bold=True, size=11, color=BLACK)
    ws.cell(row=1, column=1).fill = PatternFill('solid', fgColor='D0D8ED')
    ws.cell(row=1, column=1).alignment = Alignment(vertical='center')
    ws.cell(row=1, column=1).border = header_border
    for c in range(2, nv_cols + 1):
        cell = ws.cell(row=1, column=c, value='')
        cell.fill = PatternFill('solid', fgColor='D0D8ED')
        cell.border = header_border
    kh_cell = ws.cell(row=1, column=kh_col, value='KHÁCH HÀNG')
    kh_cell.font = Font(name=FONT_NAME, bold=True, size=11, color=BLACK)
    kh_cell.fill = PatternFill('solid', fgColor='D0D8ED')
    kh_cell.alignment = Alignment(vertical='center')
    kh_cell.border = header_border
    for ci, ch in enumerate(col_headers):
        cell = ws.cell(row=1, column=data_start + ci, value=ch)
        cell.font = Font(name=FONT_NAME, bold=True, size=11, color=BLACK)
        cell.fill = PatternFill('solid', fgColor='D0D8ED')
        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        cell.border = header_border
    ws.row_dimensions[1].height = 36

    for row_data in rows:
        cur_row += 1
        rtype = row_data.get('type', '')
        depth = row_data.get('depth', 0)
        name = row_data.get('name', '')
        values = row_data.get('values', [])
        total_cols = data_start + len(col_headers) - 1

        if rtype == 'nv':
            nv_col_idx = min(depth, nv_cols - 1) + 1
            ws.cell(row=cur_row, column=nv_col_idx, value=name)
            bg = nv_bg(depth)
            sz = nv_font_sz(depth)
            for c in range(1, total_cols + 1):
                cell = ws.cell(row=cur_row, column=c)
                cell.fill = PatternFill('solid', fgColor=bg)
                cell.border = thin_border
                if c >= data_start:
                    cell.font = Font(name=FONT_NAME, bold=True, size=sz, color=BLACK)
                    cell.alignment = Alignment(vertical='center', horizontal='right')
                    cell.number_format = '#,##0'
                else:
                    cell.font = Font(name=FONT_NAME, bold=True, size=sz, color=BLACK)
                    cell.alignment = Alignment(vertical='center')
            for vi, v in enumerate(values):
                if v is not None and v != '':
                    ws.cell(row=cur_row, column=data_start + vi, value=v)

        elif rtype == 'kh':
            ws.cell(row=cur_row, column=kh_col, value=name)
            for c in range(1, total_cols + 1):
                cell = ws.cell(row=cur_row, column=c)
                cell.fill = PatternFill('solid', fgColor='FFFFFF')
                cell.border = thin_border
                if c >= data_start:
                    cell.font = Font(name=FONT_NAME, size=10, color=BLACK)
                    cell.alignment = Alignment(vertical='center', horizontal='right')
                    cell.number_format = '#,##0'
                else:
                    cell.font = Font(name=FONT_NAME, size=10, color=BLACK)
                    cell.alignment = Alignment(vertical='center')
            for vi, v in enumerate(values):
                if v is not None and v != '':
                    ws.cell(row=cur_row, column=data_start + vi, value=v)

        elif rtype == 'total':
            ws.cell(row=cur_row, column=1, value='TỔNG CỘNG')
            for c in range(1, total_cols + 1):
                cell = ws.cell(row=cur_row, column=c)
                cell.fill = PatternFill('solid', fgColor='B8C6F0')
                cell.border = total_border
                if c >= data_start:
                    cell.font = Font(name=FONT_NAME, bold=True, size=11, color=BLACK)
                    cell.alignment = Alignment(vertical='center', horizontal='right')
                    cell.number_format = '#,##0'
                else:
                    cell.font = Font(name=FONT_NAME, bold=True, size=11, color=BLACK)
                    cell.alignment = Alignment(vertical='center')
            for vi, v in enumerate(values):
                if v is not None and v != '':
                    ws.cell(row=cur_row, column=data_start + vi, value=v)

        ws.row_dimensions[cur_row].height = 20

    for c in range(1, nv_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 6
    ws.column_dimensions[get_column_letter(kh_col)].width = 32
    for ci in range(len(col_headers)):
        ws.column_dimensions[get_column_letter(data_start + ci)].width = 22
    ws.freeze_panes = ws.cell(row=2, column=data_start)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    now = datetime.now()
    bp_tag = body.get('bp', '')
    filename = f'BaoCaoKD{"_" + bp_tag if bp_tag else ""}_{now.strftime("%Y%m%d")}.xlsx'

    set_api_result(
        status='ok',
        row_count=len(rows),
        meta={'export': filename, 'bp': bp_tag}
    )

    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)