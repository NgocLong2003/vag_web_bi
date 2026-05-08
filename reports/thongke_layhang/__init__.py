"""
Thống kê lấy hàng — Blueprint (DuckDB version)
Bảng phẳng: Dòng SP → Sản phẩm → Khu vực → Khách hàng → SL, Doanh số
"""
from flask import Blueprint, request, send_file, current_app
from api_logger import api_response, set_api_result
from datetime import datetime
import logging

logger = logging.getLogger(__name__)

bp = Blueprint('tklh', __name__,
               url_prefix='/reports/thong-ke-lay-hang',
               template_folder='templates')
bp.api_report = 'thong-ke-lay-hang'
from query_loader import load_sql


def get_store():
    return current_app.config['DUCKDB_STORE']


@bp.route('/api/ky-bao-cao')
def api_ky_bao_cao():
    try:
        data = get_store().query(load_sql('KY_BAO_CAO_DUCK'))
        return api_response(ok=True, data=data, count=len(data))
    except Exception as e:
        logger.error(f"[TKLH ky_bao_cao] {e}")
        return api_response(ok=False, error=str(e))


@bp.route('/api/hierarchy')
def api_hierarchy():
    try:
        data = get_store().query(load_sql('HIERARCHY_CTE_DUCK'))
        return api_response(ok=True, data=data, count=len(data))
    except Exception as e:
        logger.error(f"[TKLH hierarchy] {e}")
        return api_response(ok=False, error=str(e))


@bp.route('/api/khachhang')
def api_khachhang():
    try:
        data = get_store().query(load_sql('KHACHHANG_DUCK'))
        return api_response(ok=True, data=data, count=len(data))
    except Exception as e:
        logger.error(f"[TKLH khachhang] {e}")
        return api_response(ok=False, error=str(e))


@bp.route('/api/data', methods=['POST'])
def api_data():
    body = request.get_json(force=True)
    ngay_a = body.get('ngay_a', '')
    ngay_b = body.get('ngay_b', '')
    ma_bp = body.get('ma_bp', '')
    ds_nvkd = body.get('ds_nvkd', '')
    ds_kh = body.get('ds_kh', '')

    if not ngay_a or not ngay_b:
        return api_response(ok=False, error='Thiếu ngay_a hoặc ngay_b', status_code=400)

    try:
        data = get_store().query(
            load_sql('THONGKE_LAYHANG_DUCK'),
            [ngay_a, ngay_b, ma_bp or '', ds_nvkd or '', ds_kh or '']
        )
        for d in data:
            for k in ('so_luong', 'doanhso'):
                if d.get(k) is not None:
                    d[k] = float(d[k])
        return api_response(ok=True, data=data, count=len(data),
                            meta={'ngay_a': ngay_a, 'ngay_b': ngay_b})
    except Exception as e:
        logger.error(f"[TKLH data] {e}")
        return api_response(ok=False, error=str(e))


@bp.route('/api/export_excel', methods=['POST'])
def api_export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO

    body = request.get_json(force=True)
    rows = body.get('rows', [])
    kbc_name = body.get('kbc_name', '')

    if not rows:
        return api_response(ok=False, error='Không có dữ liệu', status_code=400)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Thống kê lấy hàng'
    FN = 'Arial'
    BK = '000000'
    hdr_fill = PatternFill('solid', fgColor='D0D8ED')
    hdr_font = Font(name=FN, bold=True, size=11, color=BK)
    hdr_border = Border(bottom=Side(style='medium', color='A0AAC0'))
    cell_border = Border(bottom=Side(style='thin', color='D5D9E4'))
    cell_font = Font(name=FN, size=10, color=BK)

    headers = ['Dòng SP', 'Sản phẩm', 'Khu vực', 'Khách hàng', 'SL', 'Doanh số']
    for ci, h in enumerate(headers):
        cell = ws.cell(row=1, column=ci + 1, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = hdr_border
        cell.alignment = Alignment(vertical='center',
                                   horizontal='center' if ci >= 4 else 'left')
    ws.row_dimensions[1].height = 28

    for ri, r in enumerate(rows):
        row_num = ri + 2
        ws.cell(row=row_num, column=1, value=r.get('ten_thuoc', ''))
        ws.cell(row=row_num, column=2, value=r.get('ten_vt', ''))
        ws.cell(row=row_num, column=3, value=r.get('ten_plkh3', ''))
        ws.cell(row=row_num, column=4, value=r.get('ten_kh', ''))
        c5 = ws.cell(row=row_num, column=5, value=r.get('so_luong', 0))
        c5.number_format = '#,##0'
        c6 = ws.cell(row=row_num, column=6, value=r.get('doanhso', 0))
        c6.number_format = '#,##0'
        for c in range(1, 7):
            cell = ws.cell(row=row_num, column=c)
            cell.font = cell_font
            cell.border = cell_border
            if c >= 5:
                cell.alignment = Alignment(horizontal='right')

    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 18
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:F{len(rows) + 1}'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    now = datetime.now()
    filename = f'ThongKeLayHang{"_" + kbc_name if kbc_name else ""}_{now.strftime("%Y%m%d")}.xlsx'

    set_api_result(status='ok', row_count=len(rows),
                   meta={'export': filename})

    return send_file(buf,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)