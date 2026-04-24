"""
reports/san_xuat/canhbao_tonkho/__init__.py — Cảnh báo tồn kho
Datasource: sanxuat (SQL Server realtime)
Gọi stored procedure asINRptCB_DMAT_Flat
"""
from flask import Blueprint, request, send_file
from api_logger import api_response
from datetime import date, datetime as dt
from decimal import Decimal
from io import BytesIO
import logging
import pyodbc

logger = logging.getLogger(__name__)

bp = Blueprint('cbtk', __name__, url_prefix='/reports/canh-bao-ton-kho')
bp.api_report = 'canh-bao-ton-kho'


SP_TIMEOUT = 180


def _get_conn():
    from config import DATASOURCES
    c = DATASOURCES.get('sanxuat')
    if not c:
        raise RuntimeError("DATASOURCES['sanxuat'] chưa cấu hình")
    conn = pyodbc.connect(
        f"DRIVER={{{c['driver']}}};SERVER={c['server']},{c['port']};"
        f"DATABASE={c['database']};UID={c['username']};PWD={c['password']};"
        "TrustServerCertificate=yes;Connect Timeout=30;",
        timeout=30,
        autocommit=True
    )
    conn.timeout = SP_TIMEOUT
    return conn


def _serialize(row_dict):
    for k, v in row_dict.items():
        if isinstance(v, Decimal):
            row_dict[k] = float(v)
        elif isinstance(v, dt):
            row_dict[k] = v.isoformat()
    return row_dict


KEEP_COLS = ('ma_vt', 'ten_vt', 'tam_nhap', 'ton_kho_thuc', 'sl_antoan', 'chenh_lech', 'dang_giao')
NUM_COLS = ('tam_nhap', 'ton_kho_thuc', 'sl_antoan', 'chenh_lech', 'dang_giao')


@bp.route('/api/data')
def api_data():
    ngay = request.args.get('ngay', '').strip()
    ngay_dh1 = request.args.get('ngay_dh1', '').strip()

    if not ngay:
        ngay = date.today().isoformat()
    if not ngay_dh1:
        ngay_dh1 = '2025-01-01'

    conn = None
    t0 = dt.now()
    try:
        conn = _get_conn()
        cur = conn.cursor()
        cur.execute(
            "SET NOCOUNT ON; EXEC asINRptCB_DMAT_Flat @pNgay=?, @pNgayDH1=?",
            [ngay, ngay_dh1]
        )

        while cur.description is None:
            if not cur.nextset():
                break

        if cur.description is None:
            return api_response(ok=True, rows=[], count=0,
                                meta={'ngay': ngay, 'ngay_dh1': ngay_dh1})

        columns = [d[0] for d in cur.description]
        rows_raw = cur.fetchall()

        rows = []
        for r in rows_raw:
            d = _serialize(dict(zip(columns, r)))
            out = {}
            for k in KEEP_COLS:
                out[k] = d.get(k, '' if k in ('ma_vt', 'ten_vt') else 0.0)
            for k in NUM_COLS:
                v = out.get(k)
                if v is None or v == '':
                    out[k] = 0.0
                else:
                    try:
                        out[k] = float(v)
                    except:
                        out[k] = 0.0
            rows.append(out)

        elapsed = (dt.now() - t0).total_seconds()
        logger.info(f"[cbtk] SP trả {len(rows)} dòng trong {elapsed:.1f}s (ngay={ngay}, ngay_dh1={ngay_dh1})")

        return api_response(ok=True, rows=rows,
                            elapsed_ms=int(elapsed * 1000),
                            meta={'ngay': ngay, 'ngay_dh1': ngay_dh1})

    except pyodbc.OperationalError as e:
        elapsed = (dt.now() - t0).total_seconds()
        logger.error(f'[cbtk] Timeout/OperationalError sau {elapsed:.1f}s: {e}')
        return api_response(
            ok=False,
            error=f'SP chạy quá lâu (>{SP_TIMEOUT}s) hoặc mất kết nối. Thử lại sau vài giây.',
            meta={'ngay': ngay, 'ngay_dh1': ngay_dh1}
        )
    except Exception as e:
        logger.exception('[cbtk] Data error')
        return api_response(ok=False, error=str(e),
                            meta={'ngay': ngay, 'ngay_dh1': ngay_dh1})
    finally:
        if conn:
            try:
                conn.close()
            except:
                pass


@bp.route('/api/vattu')
def api_vattu():
    try:
        from datasource import get_ds
        ds = get_ds('sanxuat')
        rows = ds.query("""
            SELECT DISTINCT ma_vt, ten_vt, dvt
            FROM [AE1213VietAnhDATA].[dbo].[DMHANGHOA_VIEW]
            ORDER BY ma_vt
        """)
        return api_response(ok=True, items=rows, count=len(rows))
    except Exception as e:
        logger.error(f'[cbtk] vattu error: {e}')
        return api_response(ok=False, error=str(e))


DG_KEEP = ('ngay_dat', 'ten_kh', 'ten_nha_sx', 'dang_giao')
DG_NUM = ('dang_giao')


@bp.route('/api/danggiao')
def api_danggiao():
    """Chi tiết đơn đang giao cho 1 mã vật tư."""
    ma_vt = request.args.get('ma_vt', '').strip()
    ngay = request.args.get('ngay', '').strip()
    ngay_dh1 = request.args.get('ngay_dh1', '').strip()

    if not ma_vt:
        return api_response(ok=False, error='Thiếu mã vật tư', status_code=400)
    if not ngay:
        ngay = date.today().isoformat()
    if not ngay_dh1:
        ngay_dh1 = '2025-01-01'

    conn = None
    t0 = dt.now()
    try:
        conn = _get_conn()
        cur = conn.cursor()
        cur.execute(
            "SET NOCOUNT ON; EXEC asINRptCB_DMAT_DangGiao @pMa_vt=?, @pNgay=?, @pNgayDH1=?",
            [ma_vt, ngay, ngay_dh1]
        )

        while cur.description is None:
            if not cur.nextset():
                break

        if cur.description is None:
            return api_response(ok=True, rows=[], count=0,
                                meta={'ma_vt': ma_vt, 'ngay': ngay, 'ngay_dh1': ngay_dh1})

        columns = [d[0] for d in cur.description]
        rows_raw = cur.fetchall()

        rows = []
        for r in rows_raw:
            d = _serialize(dict(zip(columns, r)))
            out = {}
            for k in DG_KEEP:
                out[k] = d.get(k, '')
            for k in DG_NUM:
                v = out.get(k)
                if v is None or v == '':
                    out[k] = 0.0
                else:
                    try:
                        out[k] = float(v)
                    except:
                        out[k] = 0.0
            # Format ngay_dat
            nd = out.get('ngay_dat', '')
            if nd and len(str(nd)) >= 10:
                out['ngay_dat'] = str(nd)[:10]
            rows.append(out)

        elapsed = (dt.now() - t0).total_seconds()
        logger.info(f"[cbtk] DangGiao {ma_vt}: {len(rows)} dòng trong {elapsed:.1f}s")

        return api_response(ok=True, rows=rows,
                            elapsed_ms=int(elapsed * 1000),
                            meta={'ma_vt': ma_vt, 'ngay': ngay, 'ngay_dh1': ngay_dh1})

    except Exception as e:
        logger.exception('[cbtk] DangGiao error')
        return api_response(ok=False, error=str(e),
                            meta={'ma_vt': ma_vt})
    finally:
        if conn:
            try:
                conn.close()
            except:
                pass


# ═══════════════════════════════════════════════════════════════════════════
# EXPORT EXCEL — nhận rows JSON từ client (đã được filter/sort đúng như UI)
# ═══════════════════════════════════════════════════════════════════════════
@bp.route('/api/export-excel', methods=['POST'])
def api_export_excel():
    """
    Client gửi POST JSON:
      {
        "rows": [...],          # đã filter + sort theo UI hiện tại
        "meta": {
          "ngay": "2026-04-24",
          "ngay_dh1": "2025-01-01",
          "severity": "all|danger|warn|ok|neg",
          "search": "..."
        }
      }
    Trả về file .xlsx format đẹp.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
    from openpyxl.utils import get_column_letter

    try:
        payload = request.get_json(silent=True) or {}
        rows = payload.get('rows') or []
        meta = payload.get('meta') or {}

        ngay = meta.get('ngay', '')
        ngay_dh1 = meta.get('ngay_dh1', '')
        severity = meta.get('severity', 'all')
        search = (meta.get('search') or '').strip()

        sev_label_map = {
            'all': 'Tất cả',
            'danger': 'Thiếu hàng',
            'warn': 'Cận an toàn',
            'ok': 'Đủ hàng',
            'neg': 'Tồn âm',
        }
        sev_label = sev_label_map.get(severity, severity)

        # ── Workbook setup ─────────────────────────────────────────────
        wb = Workbook()
        ws = wb.active
        ws.title = 'Cảnh báo tồn kho'

        # Palette (khớp UI)
        C_HEADER_BG = 'FF1E2A3A'   # g7
        C_HEADER_FG = 'FFFFFFFF'
        C_TITLE_FG = 'FF1E2A3A'
        C_SUBTITLE_FG = 'FF5A6478'  # g5
        C_META_LABEL = 'FF8B95AA'   # g4
        C_BORDER = 'FFD8DCE6'       # g2
        C_ZEBRA = 'FFF7F8FA'        # g0
        C_DANGER = 'FFC0392B'       # red
        C_DANGER_BG = 'FFFEF2F2'    # red-lt
        C_WARN = 'FFB45309'         # amber
        C_WARN_BG = 'FFFEF3C7'      # amber-lt
        C_OK = 'FF056944'           # emr
        C_OK_BG = 'FFEAFAF4'        # emr-lt
        C_NEG_BG = 'FFFEE2E2'       # soft red cho tồn âm

        FONT_NAME = 'Calibri'

        thin = Side(border_style='thin', color=C_BORDER)
        border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

        # ── Header block (merged title + meta) ─────────────────────────
        ws.merge_cells('A1:G1')
        c = ws['A1']
        c.value = 'BÁO CÁO CẢNH BÁO TỒN KHO'
        c.font = Font(name=FONT_NAME, size=16, bold=True, color=C_TITLE_FG)
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 28

        ws.merge_cells('A2:G2')
        c = ws['A2']
        sub = f'Ngày tồn: {ngay}   •   ĐH từ: {ngay_dh1}   •   Mức cảnh báo: {sev_label}'
        if search:
            sub += f'   •   Tìm: "{search}"'
        c.value = sub
        c.font = Font(name=FONT_NAME, size=10, italic=True, color=C_SUBTITLE_FG)
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 18

        ws.merge_cells('A3:G3')
        c = ws['A3']
        c.value = f'Xuất lúc: {dt.now().strftime("%d/%m/%Y %H:%M:%S")}   •   Tổng: {len(rows):,} dòng'
        c.font = Font(name=FONT_NAME, size=9, color=C_META_LABEL)
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[3].height = 16

        # Trống 1 dòng
        ws.row_dimensions[4].height = 6

        # ── Column headers (row 5) ─────────────────────────────────────
        HEADER_ROW = 5
        headers = [
            ('STT', 'stt', 6),
            ('Mã VT', 'ma_vt', 16),
            ('Tên vật tư', 'ten_vt', 48),
            ('Tạm nhập', 'tam_nhap', 14),
            ('Tồn thực', 'ton_kho_thuc', 14),
            ('SL an toàn', 'sl_antoan', 14),
            ('Chênh lệch', 'chenh_lech', 14),
            ('Đang giao', 'dang_giao', 14),
        ]

        # Thêm cột STT — điều chỉnh width array (A..H)
        for idx, (label, _, width) in enumerate(headers, start=1):
            col_letter = get_column_letter(idx)
            cell = ws.cell(row=HEADER_ROW, column=idx, value=label)
            cell.font = Font(name=FONT_NAME, size=11, bold=True, color=C_HEADER_FG)
            cell.fill = PatternFill('solid', fgColor=C_HEADER_BG)
            cell.alignment = Alignment(
                horizontal='center' if idx in (1,) else ('left' if idx in (2, 3) else 'right'),
                vertical='center',
                wrap_text=True,
            )
            cell.border = border_all
            ws.column_dimensions[col_letter].width = width

        ws.row_dimensions[HEADER_ROW].height = 32

        # ── Data rows ──────────────────────────────────────────────────
        num_fmt = '#,##0.####;[Red]-#,##0.####;"-"'
        int_fmt = '#,##0;[Red]-#,##0;"-"'

        def classify(r):
            """Trả về 'danger' | 'warn' | 'ok' | None và is_neg"""
            chenh = r.get('chenh_lech') or 0
            sl_at = r.get('sl_antoan') or 0
            ton = r.get('ton_kho_thuc') or 0
            is_neg = ton < 0
            if chenh < 0:
                sev = 'danger'
            elif chenh == 0 and sl_at > 0:
                sev = 'warn'
            elif chenh > 0:
                sev = 'ok'
            else:
                sev = None
            return sev, is_neg

        for i, r in enumerate(rows, start=1):
            row_excel = HEADER_ROW + i
            sev, is_neg = classify(r)

            # Row fill: tồn âm ưu tiên, rồi zebra
            row_fill = None
            if is_neg:
                row_fill = PatternFill('solid', fgColor=C_NEG_BG)
            elif i % 2 == 0:
                row_fill = PatternFill('solid', fgColor=C_ZEBRA)

            values = [
                i,
                r.get('ma_vt', ''),
                r.get('ten_vt', ''),
                r.get('tam_nhap', 0) or 0,
                r.get('ton_kho_thuc', 0) or 0,
                r.get('sl_antoan', 0) or 0,
                r.get('chenh_lech', 0) or 0,
                r.get('dang_giao', 0) or 0,
            ]

            for col_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=row_excel, column=col_idx, value=val)
                cell.font = Font(name=FONT_NAME, size=10, color='FF1E2A3A')
                cell.border = border_all
                if row_fill:
                    cell.fill = row_fill

                if col_idx == 1:  # STT
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif col_idx == 2:  # Mã VT
                    cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
                    cell.font = Font(name=FONT_NAME, size=10, color='FF5A6478')
                elif col_idx == 3:  # Tên VT
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False, indent=1)
                else:  # số
                    cell.alignment = Alignment(horizontal='right', vertical='center', indent=1)
                    cell.number_format = num_fmt

            # Highlight cột Chênh lệch theo severity
            chenh_cell = ws.cell(row=row_excel, column=7)
            if sev == 'danger':
                chenh_cell.font = Font(name=FONT_NAME, size=10, bold=True, color=C_DANGER)
                chenh_cell.fill = PatternFill('solid', fgColor=C_DANGER_BG)
            elif sev == 'ok':
                chenh_cell.font = Font(name=FONT_NAME, size=10, bold=True, color=C_OK)
                chenh_cell.fill = PatternFill('solid', fgColor=C_OK_BG)
            elif sev == 'warn':
                chenh_cell.font = Font(name=FONT_NAME, size=10, bold=True, color=C_WARN)
                chenh_cell.fill = PatternFill('solid', fgColor=C_WARN_BG)

            # Tồn kho thực âm — in đỏ
            ton_cell = ws.cell(row=row_excel, column=5)
            if (r.get('ton_kho_thuc') or 0) < 0:
                ton_cell.font = Font(name=FONT_NAME, size=10, bold=True, color=C_DANGER)

            # Row height
            ws.row_dimensions[row_excel].height = 18

        # ── Freeze panes + filter ──────────────────────────────────────
        ws.freeze_panes = ws.cell(row=HEADER_ROW + 1, column=1)

        last_row = HEADER_ROW + len(rows)
        if len(rows) > 0:
            ws.auto_filter.ref = f'A{HEADER_ROW}:H{last_row}'

        # ── Page setup ─────────────────────────────────────────────────
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_margins.left = 0.4
        ws.page_margins.right = 0.4
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5
        ws.print_title_rows = f'{HEADER_ROW}:{HEADER_ROW}'

        # ── Save to BytesIO ────────────────────────────────────────────
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        # Filename
        fname_date = (ngay or dt.now().strftime('%Y-%m-%d')).replace('-', '')
        fname = f'CanhBaoTonKho_{fname_date}_{dt.now().strftime("%H%M%S")}.xlsx'

        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=fname,
        )

    except Exception as e:
        logger.exception('[cbtk] Export Excel error')
        return api_response(ok=False, error=f'Lỗi xuất Excel: {e}', status_code=500)