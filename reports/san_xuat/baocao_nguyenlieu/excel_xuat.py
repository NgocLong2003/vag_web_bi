"""
reports/san_xuat/baocao_nguyenlieu/excel_xuat.py
Tạo file Excel "THỐNG KÊ XUẤT NGUYÊN LIỆU" — layout A4 ngang.
Hỗ trợ thêm bảng thành phẩm xuất bán bên dưới.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime, timezone, timedelta
import io

# ── Shared constants ──
FN = "Cambria"
SZ = 13
VN_TZ = timezone(timedelta(hours=7))


def _styles():
    """Return dict of reusable styles."""
    thin = Side(style="thin")
    return {
        'f_title': Font(name=FN, size=16, bold=True),
        'f_lbl': Font(name=FN, size=SZ, bold=True),
        'f_val': Font(name=FN, size=SZ),
        'f_col': Font(name=FN, size=SZ, bold=True),
        'f_data': Font(name=FN, size=SZ),
        'f_total': Font(name=FN, size=SZ, bold=True),
        'f_avg': Font(name=FN, size=SZ, bold=True, italic=True),
        'f_marker': Font(name=FN, size=SZ, bold=True),
        'a_c': Alignment(horizontal="center", vertical="center"),
        'a_l': Alignment(horizontal="left", vertical="center"),
        'a_r': Alignment(horizontal="right", vertical="center"),
        'bdr': Border(left=thin, right=thin, top=thin, bottom=thin),
        'fill_h': PatternFill("solid", fgColor="D0D0D0"),
        'fill_t': PatternFill("solid", fgColor="D9D9D9"),
        'fill_avg': PatternFill("solid", fgColor="E6E6E6"),
        'fill_stripe': PatternFill("solid", fgColor="E8E8E8"),
    }


def _write_pivot_table(ws, s, ts, years, data, now_m, now_y):
    """
    Ghi bảng pivot 12 tháng × years bắt đầu từ row ts.
    data: dict[month][year] = value
    Trả về row cuối cùng (row của Trung bình).
    """
    num_yr = len(years)

    # Header row
    cell = ws.cell(row=ts, column=1, value="Tháng")
    cell.font = s['f_col']; cell.alignment = s['a_c']; cell.border = s['bdr']; cell.fill = s['fill_h']
    for i, y in enumerate(years):
        cell = ws.cell(row=ts, column=i + 2, value=y)
        cell.font = s['f_col']; cell.alignment = s['a_c']; cell.border = s['bdr']; cell.fill = s['fill_h']
        cell.number_format = "0"

    # Data rows (12 months)
    for m in range(1, 13):
        r = ts + m
        is_stripe = (m % 2 == 0)
        cell = ws.cell(row=r, column=1, value=m)
        cell.font = s['f_data']; cell.alignment = s['a_c']; cell.border = s['bdr']
        if is_stripe:
            cell.fill = s['fill_stripe']
        for i, y in enumerate(years):
            val = data.get(m, {}).get(y, 0)
            cell = ws.cell(row=r, column=i + 2)
            cell.value = val if val else 0
            cell.number_format = "#,##0"
            cell.font = s['f_data']; cell.alignment = s['a_r']; cell.border = s['bdr']
            if is_stripe:
                cell.fill = s['fill_stripe']

    # Tổng
    rt = ts + 13
    cell = ws.cell(row=rt, column=1, value="Tổng")
    cell.font = s['f_total']; cell.alignment = s['a_c']; cell.border = s['bdr']; cell.fill = s['fill_t']
    for i in range(num_yr):
        cl = get_column_letter(i + 2)
        cell = ws.cell(row=rt, column=i + 2)
        cell.value = f"=SUM({cl}{ts+1}:{cl}{ts+12})"
        cell.number_format = "#,##0"
        cell.font = s['f_total']; cell.alignment = s['a_r']; cell.border = s['bdr']; cell.fill = s['fill_t']

    # Trung bình: năm đã qua ÷12, năm hiện tại ÷ tháng hiện tại
    ra = rt + 1
    cell = ws.cell(row=ra, column=1, value="Trung bình")
    cell.font = s['f_avg']; cell.alignment = s['a_c']; cell.border = s['bdr']; cell.fill = s['fill_avg']
    for i, y in enumerate(years):
        cl = get_column_letter(i + 2)
        if y < now_y:
            divisor = 12
        elif y == now_y:
            divisor = now_m
        else:
            divisor = 1  # năm tương lai, tránh chia 0
        cell = ws.cell(row=ra, column=i + 2)
        cell.value = f"=IF({cl}{rt}=0,0,{cl}{rt}/{divisor})"
        cell.number_format = "#,##0"
        cell.font = s['f_avg']; cell.alignment = s['a_r']; cell.border = s['bdr']; cell.fill = s['fill_avg']

    return ra


def build_xuat_excel(rows, ma_vt, ten_vt, dvt, incl_xuatle=False, products=None):
    """
    rows: NVL data từ sp_TonKho_ChiTiet_TheoThang
    products: list of dict, mỗi item = {ma_vt, ten_vt, dvt, ton, rows: [...]}
              rows bên trong mỗi product có cấu trúc {thang, nam, XUAT, ...}
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Xuất NVL"
    s = _styles()

    # ── Page setup ──
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "landscape"
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.6
    ws.page_margins.bottom = 0.5
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0  # multi-page OK

    # ── Years ──
    year_set = set()
    for r in rows:
        year_set.add(int(r["nam"]))
    years = sorted(year_set)
    while len(years) < 9:
        years.append(years[-1] + 1)
    num_yr = len(years)
    last_col = max(num_yr + 1, 10)

    # ── Column widths ──
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 50
    for i in range(num_yr):
        ws.column_dimensions[get_column_letter(i + 2)].width = 12

    # ══════════════════════════════════════════
    # PHẦN 1: NGUYÊN LIỆU
    # ══════════════════════════════════════════

    # Row 1: Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    c = ws.cell(row=1, column=1, value="THỐNG KÊ XUẤT NGUYÊN LIỆU")
    c.font = s['f_title']; c.alignment = s['a_c']

    # Rows 3-5: Info trái
    ws.cell(row=3, column=1, value="Mã hàng").font = s['f_lbl']
    ws.cell(row=3, column=1).alignment = s['a_l']
    ws.cell(row=3, column=2, value=ma_vt).font = s['f_val']
    ws.cell(row=3, column=2).alignment = s['a_l']

    ws.cell(row=4, column=1, value="Tên hàng").font = s['f_lbl']
    ws.cell(row=4, column=1).alignment = s['a_l']
    ws.cell(row=4, column=2, value=ten_vt).font = s['f_val']
    ws.cell(row=4, column=2).alignment = s['a_l']

    ws.cell(row=5, column=1, value="Đơn vị tính").font = s['f_lbl']
    ws.cell(row=5, column=1).alignment = s['a_l']
    ws.cell(row=5, column=2, value=dvt).font = s['f_val']
    ws.cell(row=5, column=2).alignment = s['a_l']

    # Rows 3-4 phải: marker info
    ton_cuoiki, latest_ym = 0, 0
    for rec in rows:
        ym = int(rec["nam"]) * 100 + int(rec["thang"])
        if ym > latest_ym:
            latest_ym = ym
            ton_cuoiki = float(rec.get("CUOI_KI", 0) or 0)

    now_vn = datetime.now(VN_TZ)
    now_str = now_vn.strftime("%d/%m/%Y %H:%M")
    now_m = now_vn.month
    now_y = now_vn.year

    for mr in [3, 4]:
        ws.merge_cells(start_row=mr, start_column=9, end_row=mr, end_column=10)

    ws.cell(row=3, column=8, value="Ngày xuất").font = s['f_marker']
    ws.cell(row=3, column=9, value=f": {now_str}").font = s['f_val']
    ws.cell(row=4, column=8, value="Tồn").font = s['f_marker']
    ws.cell(row=4, column=9, value=f": {int(ton_cuoiki):,} {dvt}").font = s['f_val']

    # Build NVL data dict
    nvl_data = {}
    for rec in rows:
        m, y = int(rec["thang"]), int(rec["nam"])
        if m not in nvl_data:
            nvl_data[m] = {}
        xv = float(rec.get("XUAT", 0) or 0)
        if incl_xuatle:
            xv += float(rec.get("XUAT_LE", 0) or 0)
        nvl_data[m][y] = nvl_data[m].get(y, 0) + xv

    # Write NVL table starting row 7
    last_row = _write_pivot_table(ws, s, 7, years, nvl_data, now_m, now_y)

    # ══════════════════════════════════════════
    # PHẦN 2: THÀNH PHẨM XUẤT BÁN
    # ══════════════════════════════════════════
    if products:
        # Section title — 2 rows gap
        cur = last_row + 2
        ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=last_col)
        c = ws.cell(row=cur, column=1, value="THÀNH PHẨM XUẤT BÁN")
        c.font = s['f_title']; c.alignment = s['a_c']

        for prod in products:
            p_ma = prod.get('ma_vt', '')
            p_ten = prod.get('ten_vt', '')
            p_dvt = prod.get('dvt', '')
            p_ton = prod.get('ton', 0)
            p_rows = prod.get('rows', [])

            # Info block — 2 dòng: labels rồi values
            cur += 2
            r_lbl = cur      # dòng label
            r_val = cur + 1  # dòng value

            # A: Mã thành phẩm
            ws.cell(row=r_lbl, column=1, value="Mã thành phẩm").font = s['f_lbl']
            ws.cell(row=r_lbl, column=1).alignment = s['a_l']
            ws.cell(row=r_val, column=1, value=p_ma).font = s['f_val']
            ws.cell(row=r_val, column=1).alignment = s['a_l']

            # C: Tên thành phẩm
            ws.cell(row=r_lbl, column=3, value="Tên thành phẩm").font = s['f_lbl']
            ws.cell(row=r_lbl, column=3).alignment = s['a_l']
            ws.merge_cells(start_row=r_val, start_column=3, end_row=r_val, end_column=4)
            ws.cell(row=r_val, column=3, value=p_ten).font = s['f_val']
            ws.cell(row=r_val, column=3).alignment = s['a_l']

            # F: ĐVT label, G: ĐVT value (cùng dòng label)
            ws.cell(row=r_lbl, column=6, value="ĐVT").font = s['f_lbl']
            ws.cell(row=r_lbl, column=6).alignment = s['a_l']
            ws.cell(row=r_lbl, column=7, value=p_dvt).font = s['f_val']
            ws.cell(row=r_lbl, column=7).alignment = s['a_l']

            # H: Tồn, I-J: value (merged)
            ws.cell(row=r_lbl, column=8, value="Tồn").font = s['f_marker']
            ws.cell(row=r_lbl, column=8).alignment = s['a_l']
            ws.merge_cells(start_row=r_lbl, start_column=9, end_row=r_lbl, end_column=10)
            ws.cell(row=r_lbl, column=9, value=f": {int(p_ton):,} {p_dvt}").font = s['f_val']
            ws.cell(row=r_lbl, column=9).alignment = s['a_l']

            # Build product xuat data
            p_data = {}
            for rec in p_rows:
                m, y = int(rec["thang"]), int(rec["nam"])
                if m not in p_data:
                    p_data[m] = {}
                xv = float(rec.get("XUAT", 0) or 0)
                p_data[m][y] = p_data[m].get(y, 0) + xv

            # Write table — 1 row gap after info
            tbl_start = r_val + 1
            last_row = _write_pivot_table(ws, s, tbl_start, years, p_data, now_m, now_y)
            cur = last_row

    # ── Print area ──
    lpc = get_column_letter(last_col)
    ws.print_area = f"A1:{lpc}{last_row}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf