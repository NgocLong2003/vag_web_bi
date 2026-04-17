"""
Báo cáo Khách Hàng — Blueprint (DuckDB version)
API prefix: /reports/bao-cao-khach-hang/api/...
"""
from flask import Blueprint, request, send_file, current_app
from api_logger import api_response, set_api_result
from datetime import datetime, date
import logging

logger = logging.getLogger(__name__)

bp = Blueprint('bckh', __name__,
               url_prefix='/reports/bao-cao-khach-hang',
               template_folder='templates')
bp.api_report = 'bao-cao-khach-hang'
from query_loader import load_sql


def get_store():
    return current_app.config['DUCKDB_STORE']


# ─────────────────────────────────────────
# API: Kỳ báo cáo
# ─────────────────────────────────────────
@bp.route('/api/ky-bao-cao')
def api_ky_bao_cao():
    try:
        data = get_store().query(load_sql('KY_BAO_CAO_DUCK'))
        return api_response(ok=True, data=data, count=len(data))
    except Exception as e:
        logger.error(f"[ky_bao_cao] {e}")
        return api_response(ok=False, error=str(e))


# ─────────────────────────────────────────
# API: Hierarchy
# ─────────────────────────────────────────
@bp.route('/api/hierarchy')
def api_hierarchy():
    try:
        data = get_store().query(load_sql('HIERARCHY_CTE_DUCK'))
        return api_response(ok=True, data=data, count=len(data))
    except Exception as e:
        logger.error(f"[hierarchy] {e}")
        return api_response(ok=False, error=str(e))


# ─────────────────────────────────────────
# API: Khách hàng
# ─────────────────────────────────────────
@bp.route('/api/khachhang')
def api_khachhang():
    try:
        data = get_store().query(load_sql('KHACHHANG_DUCK'))
        return api_response(ok=True, data=data, count=len(data))
    except Exception as e:
        logger.error(f"[khachhang] {e}")
        return api_response(ok=False, error=str(e))


# ─────────────────────────────────────────
# API: Công nợ
# ─────────────────────────────────────────
@bp.route('/api/congno', methods=['POST'])
def api_congno():
    body = request.get_json(force=True)
    ngay_cut = body.get('ngay_cut', '')
    ma_bp = body.get('ma_bp', '')
    ds_nvkd = body.get('ds_nvkd', '')
    ds_kh = body.get('ds_kh', '')

    if not ngay_cut:
        return api_response(ok=False, error='Thiếu ngay_cut', status_code=400)
    try:
        start_y = datetime.strptime(ngay_cut, '%Y-%m-%d').year
    except ValueError:
        return api_response(ok=False, error='ngay_cut không hợp lệ', status_code=400)

    try:
        data = get_store().query(
            load_sql('CONGNO_SQL_DUCK'),
            [ngay_cut, start_y, ma_bp or '', ds_nvkd or '', ds_kh or '']
        )
        for d in data:
            for k in ('so_du_ban_dau', 'tong_phatsinh', 'du_no_ck'):
                if d.get(k) is not None:
                    d[k] = float(d[k])
        return api_response(ok=True, data=data, count=len(data),
                            meta={'ngay_cut': ngay_cut, 'ma_bp': ma_bp})
    except Exception as e:
        logger.error(f"[BCKH congno] {e}")
        return api_response(ok=False, error=str(e))


# ─────────────────────────────────────────
# API: Doanh số (bán ra)
# ─────────────────────────────────────────
@bp.route('/api/doanhso', methods=['POST'])
def api_doanhso():
    body = request.get_json(force=True)
    ngay_a = body.get('ngay_a', '')
    ngay_b = body.get('ngay_b', '')
    ma_bp = body.get('ma_bp', '')
    ds_nvkd = body.get('ds_nvkd', '')
    ds_kh = body.get('ds_kh', '')

    if not ngay_a or not ngay_b:
        return api_response(ok=False, error='Thiếu ngay_a hoặc ngay_b', status_code=400)

    try:
        data = get_store().query(
            load_sql('DOANHSO_SQL_DUCK'),
            [ngay_a, ngay_b, ma_bp or '', ds_nvkd or '', ds_kh or '']
        )
        for d in data:
            for k in ('tong_so_luong', 'tong_tien_nt2', 'tong_tien_ck_nt', 'tong_thue_gtgt_nt', 'tong_doanhso'):
                if d.get(k) is not None:
                    d[k] = float(d[k])
        return api_response(ok=True, data=data, count=len(data),
                            meta={'ngay_a': ngay_a, 'ngay_b': ngay_b, 'ma_bp': ma_bp})
    except Exception as e:
        logger.error(f"[BCKH doanhso] {e}")
        return api_response(ok=False, error=str(e))


# ─────────────────────────────────────────
# API: Doanh thu (thanh toán) — 2 luồng ngày
# ─────────────────────────────────────────
@bp.route('/api/doanhthu', methods=['POST'])
def api_doanhthu():
    body = request.get_json(force=True)
    ngay_a = body.get('ngay_a', '')
    ngay_b = body.get('ngay_b', '')
    ngay_a2 = body.get('ngay_a2', '')
    ngay_b2 = body.get('ngay_b2', '')
    ma_bp = body.get('ma_bp', '')
    ds_nvkd = body.get('ds_nvkd', '')
    ds_kh = body.get('ds_kh', '')

    if not ngay_a or not ngay_b:
        return api_response(ok=False, error='Thiếu ngay_a hoặc ngay_b', status_code=400)

    try:
        data = get_store().query(
            load_sql('DOANHTHU_BCKH_DUCK'),
            [ngay_a, ngay_b, ngay_a2 or None, ngay_b2 or None,
             ma_bp or '', ds_nvkd or '', ds_kh or '']
        )
        for d in data:
            if d.get('doanhthu') is not None:
                d['doanhthu'] = float(d['doanhthu'])
        return api_response(ok=True, data=data, count=len(data),
                            meta={'ngay_a': ngay_a, 'ngay_b': ngay_b, 'ma_bp': ma_bp})
    except Exception as e:
        logger.error(f"[BCKH doanhthu] {e}")
        return api_response(ok=False, error=str(e))


# ─────────────────────────────────────────
# API: Dư nợ trong kỳ
# ─────────────────────────────────────────
@bp.route('/api/dunotrongky', methods=['POST'])
def api_dunotrongky():
    body = request.get_json(force=True)
    ngay_a_hang = body.get('ngay_a_hang', '')
    ngay_b_hang = body.get('ngay_b_hang', '')
    ngay_a_tien = body.get('ngay_a_tien', '')
    ngay_b_tien = body.get('ngay_b_tien', '')
    ma_bp = body.get('ma_bp', '')
    ds_nvkd = body.get('ds_nvkd', '')
    ds_kh = body.get('ds_kh', '')

    if not ngay_a_hang or not ngay_b_hang or not ngay_a_tien or not ngay_b_tien:
        return api_response(ok=False, error='Thiếu ngày', status_code=400)

    try:
        data = get_store().query(
            load_sql('DUNOTRONGKY_DUCK'),
            [ngay_a_hang, ngay_b_hang, ngay_a_tien, ngay_b_tien,
             ma_bp or '', ds_nvkd or '', ds_kh or '']
        )
        for d in data:
            for k in ('ban_ra', 'tra_ve', 'dt_thuong', 'du_no_trong_ky'):
                if d.get(k) is not None:
                    d[k] = float(d[k])
        return api_response(ok=True, data=data, count=len(data),
                            meta={'ma_bp': ma_bp})
    except Exception as e:
        logger.error(f"[BCKH dunotrongky] {e}")
        return api_response(ok=False, error=str(e))


# ─────────────────────────────────────────
# API: Dư nợ cuối kỳ
# ─────────────────────────────────────────
@bp.route('/api/dunocuoiky', methods=['POST'])
def api_dunocuoiky():
    body = request.get_json(force=True)
    ngay_cut = body.get('ngay_cut', '')
    ngay_a_lk = body.get('ngay_a_lk', '')
    ngay_b_lk = body.get('ngay_b_lk', '')
    ma_bp = body.get('ma_bp', '')
    ds_nvkd = body.get('ds_nvkd', '')
    ds_kh = body.get('ds_kh', '')

    if not ngay_cut:
        return api_response(ok=False, error='Thiếu ngay_cut', status_code=400)
    try:
        start_y = datetime.strptime(ngay_cut, '%Y-%m-%d').year
    except ValueError:
        return api_response(ok=False, error='ngay_cut không hợp lệ', status_code=400)

    has_lk = 1 if (ngay_a_lk and ngay_b_lk) else 0

    try:
        data = get_store().query(
            load_sql('DUNOCUOIKY_DUCK'),
            [ngay_cut, start_y, ngay_a_lk or None, ngay_b_lk or None,
             has_lk, ma_bp or '', ds_nvkd or '', ds_kh or '']
        )
        for d in data:
            for k in ('du_no', 'ban_ra_lk', 'tra_ve_lk', 'du_no_cuoi_ky'):
                if d.get(k) is not None:
                    d[k] = float(d[k])
        return api_response(ok=True, data=data, count=len(data),
                            meta={'ngay_cut': ngay_cut, 'ma_bp': ma_bp})
    except Exception as e:
        logger.error(f"[BCKH dunocuoiky] {e}")
        return api_response(ok=False, error=str(e))


# ─────────────────────────────────────────
# API: Chi tiết doanh số (modal lịch sử)
# ─────────────────────────────────────────
@bp.route('/api/doanhso_chitiet', methods=['POST'])
def api_doanhso_chitiet():
    body = request.get_json(force=True)
    ngay_a = body.get('ngay_a', '')
    ngay_b = body.get('ngay_b', '')
    ma_kh = body.get('ma_kh', '')
    if not ngay_a or not ngay_b or not ma_kh:
        return api_response(ok=False, error='Thiếu tham số', status_code=400)

    try:
        data = get_store().query(
            load_sql('DOANHSO_CHITIET_BCKH_DUCK'),
            [ngay_a, ngay_b, ma_kh]
        )
        for d in data:
            for k in ('so_luong', 'gia_nt2', 'tien_nt2', 'tien_ck_nt', 'thue_gtgt_nt', 'doanhso'):
                if d.get(k) is not None:
                    d[k] = float(d[k])
        return api_response(ok=True, data=data, count=len(data),
                            meta={'ma_kh': ma_kh})
    except Exception as e:
        logger.error(f"[BCKH ds_chitiet] {e}")
        return api_response(ok=False, error=str(e))


# ─────────────────────────────────────────
# API: Chi tiết doanh thu (modal lịch sử)
# ─────────────────────────────────────────
@bp.route('/api/doanhthu_chitiet', methods=['POST'])
def api_doanhthu_chitiet():
    body = request.get_json(force=True)
    ngay_a = body.get('ngay_a', '')
    ngay_b = body.get('ngay_b', '')
    ma_kh = body.get('ma_kh', '')
    if not ngay_a or not ngay_b or not ma_kh:
        return api_response(ok=False, error='Thiếu tham số', status_code=400)

    try:
        data = get_store().query(
            load_sql('DOANHTHU_CHITIET_BCKH_DUCK'),
            [ngay_a, ngay_b, ma_kh]
        )
        for d in data:
            if d.get('doanhthu') is not None:
                d['doanhthu'] = float(d['doanhthu'])
        return api_response(ok=True, data=data, count=len(data),
                            meta={'ma_kh': ma_kh})
    except Exception as e:
        logger.error(f"[BCKH dt_chitiet] {e}")
        return api_response(ok=False, error=str(e))


# ─────────────────────────────────────────
# API: Chi tiết thưởng (modal lịch sử)
# ─────────────────────────────────────────
@bp.route('/api/thuong_chitiet', methods=['POST'])
def api_thuong_chitiet():
    body = request.get_json(force=True)
    ngay_a = body.get('ngay_a', '')
    ngay_b = body.get('ngay_b', '')
    ma_kh = body.get('ma_kh', '')
    if not ngay_a or not ngay_b or not ma_kh:
        return api_response(ok=False, error='Thiếu tham số', status_code=400)

    try:
        data = get_store().query(
            load_sql('THUONG_CHITIET_DUCK'),
            [ngay_a, ngay_b, ma_kh]
        )
        for d in data:
            if d.get('thuong') is not None:
                d['thuong'] = float(d['thuong'])
        return api_response(ok=True, data=data, count=len(data),
                            meta={'ma_kh': ma_kh})
    except Exception as e:
        logger.error(f"[BCKH thuong_chitiet] {e}")
        return api_response(ok=False, error=str(e))


# ─────────────────────────────────────────
# API: Chi tiết trả lại (modal lịch sử)
# ─────────────────────────────────────────
@bp.route('/api/tralai_chitiet', methods=['POST'])
def api_tralai_chitiet():
    body = request.get_json(force=True)
    ngay_a = body.get('ngay_a', '')
    ngay_b = body.get('ngay_b', '')
    ma_kh = body.get('ma_kh', '')
    if not ngay_a or not ngay_b or not ma_kh:
        return api_response(ok=False, error='Thiếu tham số', status_code=400)

    try:
        data = get_store().query(
            load_sql('TRALAI_CHITIET_DUCK'),
            [ngay_a, ngay_b, ma_kh]
        )
        for d in data:
            for k in ('so_luong', 'gia_nt2', 'tien_nt2', 'tien_ck_nt', 'thue_gtgt_nt', 'tralai'):
                if d.get(k) is not None:
                    d[k] = float(d[k])
        return api_response(ok=True, data=data, count=len(data),
                            meta={'ma_kh': ma_kh})
    except Exception as e:
        logger.error(f"[BCKH tralai_chitiet] {e}")
        return api_response(ok=False, error=str(e))


# ─────────────────────────────────────────
# API: Export Excel
# ─────────────────────────────────────────
@bp.route('/api/export_excel', methods=['POST'])
def api_export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    body = request.get_json(force=True)
    rows = body.get('rows', [])
    col_headers = body.get('col_headers', [])
    kbc_name = body.get('kbc_name', '')

    if not rows:
        return api_response(ok=False, error='Không có dữ liệu', status_code=400)

    max_nv_depth = max((r.get('depth', 0) for r in rows if r.get('type') == 'nv'), default=0)
    nv_cols = max_nv_depth + 1
    kh_col = nv_cols + 1
    data_start = nv_cols + 2

    wb = Workbook()
    ws = wb.active
    ws.title = 'Báo cáo KH'
    FN = 'Arial'
    BK = '000000'
    thin = Border(bottom=Side(style='thin', color='D5D9E4'), right=Side(style='thin', color='ECEEF3'))
    hdr = Border(bottom=Side(style='medium', color='A0AAC0'))
    tot = Border(top=Side(style='medium', color='8090B0'), bottom=Side(style='medium', color='8090B0'))
    bgs = ['B8C6F0', 'CADAF6', 'DAEAFC', 'E8F0FD', 'F0F5FE', 'F7FAFE']

    def nbg(d): return bgs[min(d, len(bgs) - 1)]
    def nsz(d): return 11 if d == 0 else 10.5 if d == 1 else 10

    cur_row = 1
    for c in range(1, nv_cols + 1):
        cell = ws.cell(row=1, column=c, value='NHÂN VIÊN KINH DOANH' if c == 1 else '')
        cell.font = Font(name=FN, bold=True, size=11, color=BK)
        cell.fill = PatternFill('solid', fgColor='D0D8ED')
        cell.alignment = Alignment(vertical='center')
        cell.border = hdr
    kh = ws.cell(row=1, column=kh_col, value='KHÁCH HÀNG')
    kh.font = Font(name=FN, bold=True, size=11, color=BK)
    kh.fill = PatternFill('solid', fgColor='D0D8ED')
    kh.alignment = Alignment(vertical='center')
    kh.border = hdr
    for ci, ch in enumerate(col_headers):
        cell = ws.cell(row=1, column=data_start + ci, value=ch)
        cell.font = Font(name=FN, bold=True, size=11, color=BK)
        cell.fill = PatternFill('solid', fgColor='D0D8ED')
        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        cell.border = hdr
    ws.row_dimensions[1].height = 42

    for rd in rows:
        cur_row += 1
        rt = rd.get('type', '')
        dp = rd.get('depth', 0)
        nm = rd.get('name', '')
        vs = rd.get('values', [])
        tc = data_start + len(col_headers) - 1

        if rt == 'nv':
            ws.cell(row=cur_row, column=min(dp, nv_cols - 1) + 1, value=nm)
            bg = nbg(dp)
            sz = nsz(dp)
            for c in range(1, tc + 1):
                cell = ws.cell(row=cur_row, column=c)
                cell.fill = PatternFill('solid', fgColor=bg)
                cell.border = thin
                cell.font = Font(name=FN, bold=True, size=sz, color=BK)
                cell.alignment = Alignment(vertical='center', horizontal='right' if c >= data_start else 'left')
                if c >= data_start:
                    cell.number_format = '#,##0'
            for vi, v in enumerate(vs):
                if v is not None and v != '':
                    ws.cell(row=cur_row, column=data_start + vi, value=v)

        elif rt == 'kh':
            ws.cell(row=cur_row, column=kh_col, value=nm)
            for c in range(1, tc + 1):
                cell = ws.cell(row=cur_row, column=c)
                cell.fill = PatternFill('solid', fgColor='FFFFFF')
                cell.border = thin
                cell.font = Font(name=FN, size=10, color=BK)
                cell.alignment = Alignment(vertical='center', horizontal='right' if c >= data_start else 'left')
                if c >= data_start:
                    cell.number_format = '#,##0'
            for vi, v in enumerate(vs):
                if v is not None and v != '':
                    ws.cell(row=cur_row, column=data_start + vi, value=v)

        elif rt == 'total':
            ws.cell(row=cur_row, column=1, value='TỔNG CỘNG')
            for c in range(1, tc + 1):
                cell = ws.cell(row=cur_row, column=c)
                cell.fill = PatternFill('solid', fgColor='B8C6F0')
                cell.border = tot
                cell.font = Font(name=FN, bold=True, size=11, color=BK)
                cell.alignment = Alignment(vertical='center', horizontal='right' if c >= data_start else 'left')
                if c >= data_start:
                    cell.number_format = '#,##0'
            for vi, v in enumerate(vs):
                if v is not None and v != '':
                    ws.cell(row=cur_row, column=data_start + vi, value=v)

        ws.row_dimensions[cur_row].height = 20

    for c in range(1, nv_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 6
    ws.column_dimensions[get_column_letter(kh_col)].width = 32
    for ci in range(len(col_headers)):
        ws.column_dimensions[get_column_letter(data_start + ci)].width = 22
    ws.freeze_panes = ws.cell(row=2, column=data_start)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    now = datetime.now()
    filename = f'BaoCaoKH{"_" + kbc_name if kbc_name else ""}_{now.strftime("%Y%m%d")}.xlsx'

    set_api_result(
        status='ok',
        row_count=len(rows),
        meta={'export': filename, 'kbc_name': kbc_name}
    )

    return send_file(buf,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)