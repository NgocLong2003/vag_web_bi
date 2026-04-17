"""
Báo cáo Bán Ra — Blueprint
API prefix: /reports/bao-cao-ban-ra/api/...
Query trực tiếp SQL Server (dữ liệu lớn, flat table)
"""
from flask import Blueprint, request, send_file
from api_logger import api_response, set_api_result
from datetime import datetime
import logging, io

logger = logging.getLogger(__name__)

bp = Blueprint('bcbr', __name__,
               url_prefix='/reports/bao-cao-ban-ra',
               template_folder='templates')
bp.api_report = 'bao-cao-ban-ra'

def _get_sqlserver():
    """Lấy SQL Server connection từ pyodbc."""
    from config import SQLSERVER_CONFIG
    import pyodbc
    c = SQLSERVER_CONFIG
    return pyodbc.connect(
        f"DRIVER={{{c['driver']}}};SERVER={c['server']},{c['port']};"
        f"DATABASE={c['database']};UID={c['username']};PWD={c['password']};"
        "TrustServerCertificate=yes;Connect Timeout=30;",
        autocommit=True)


def _query(sql, params=None):
    """Execute query, return list of dicts."""
    conn = _get_sqlserver()
    try:
        cur = conn.cursor()
        cur.execute(sql, params or [])
        cols = [d[0] for d in cur.description]
        rows = []
        for r in cur.fetchall():
            rows.append({cols[i]: (r[i].isoformat() if isinstance(r[i], (datetime,)) else
                                   str(r[i])[:10] if hasattr(r[i], 'isoformat') else r[i])
                         for i in range(len(cols))})
        return rows
    finally:
        conn.close()


# ─────────────────────────────────────────
# API: Dữ liệu bán ra
# ─────────────────────────────────────────
@bp.route('/api/data', methods=['POST'])
def api_data():
    try:
        body = request.get_json(force=True)
        ngay_a = body.get('ngay_a', '')
        ngay_b = body.get('ngay_b', '')
        ma_bp = body.get('ma_bp', '')
        ds_nvkd = body.get('ds_nvkd', '')
        ds_kh = body.get('ds_kh', '')
        ds_vt = body.get('ds_vt', '')
        ds_kho = body.get('ds_kho', '')
        ds_ten_bp = body.get('ds_ten_bp', '')

        if not ngay_a or not ngay_b:
            return api_response(ok=False, error='Thiếu ngày', status_code=400)

        conditions = ["ngay_ct >= ?", "ngay_ct <= ?", "ma_bp != 'TN'"]
        params = [ngay_a, ngay_b]

        if ma_bp:
            bps = [b.strip() for b in ma_bp.split(',') if b.strip()]
            if bps:
                placeholders = ','.join(['?'] * len(bps))
                conditions.append(f'ma_bp IN ({placeholders})')
                params.extend(bps)

        if ds_nvkd:
            nvs = [n.strip() for n in ds_nvkd.split(',') if n.strip()]
            if nvs:
                placeholders = ','.join(['?'] * len(nvs))
                conditions.append(f'ma_nvkd IN ({placeholders})')
                params.extend(nvs)

        if ds_kh:
            khs = [k.strip() for k in ds_kh.split(',') if k.strip()]
            if khs:
                placeholders = ','.join(['?'] * len(khs))
                conditions.append(f'ma_kh IN ({placeholders})')
                params.extend(khs)

        if ds_vt:
            vts = [v.strip() for v in ds_vt.split(',') if v.strip()]
            if vts:
                placeholders = ','.join(['?'] * len(vts))
                conditions.append(f'ma_vt IN ({placeholders})')
                params.extend(vts)

        if ds_kho:
            khos = [k.strip() for k in ds_kho.split(',') if k.strip()]
            if khos:
                placeholders = ','.join(['?'] * len(khos))
                conditions.append(f'ma_kho IN ({placeholders})')
                params.extend(khos)

        if ds_ten_bp:
            tbps = [b.strip() for b in ds_ten_bp.split(',') if b.strip()]
            if tbps:
                placeholders = ','.join(['?'] * len(tbps))
                conditions.append(f'ten_bp IN ({placeholders})')
                params.extend(tbps)

        where = ' AND '.join(conditions)
        sql = f"""
            SELECT ngay_ct, ma_kh, ten_kh_vat, dien_giai, ma_vt, ten_vt, dvt,
                   so_luong, gia_nt2, tien_nt2, tl_ck, tien_ck_nt,
                   ts_gtgt, thue_gtgt_nt, tt_nt, ma_kho, ten_bp, ma_nvkd, ten_nvkd
            FROM [dbo].[BKHDBANHANG_VIEW]
            WHERE {where}
            ORDER BY ngay_ct, ma_nvkd, ma_kh
        """
        data = _query(sql, params)
        return api_response(ok=True, data=data, count=len(data),
                            meta={'ngay_a': ngay_a, 'ngay_b': ngay_b, 'ma_bp': ma_bp})
    except Exception as e:
        logger.exception('api_data error')
        return api_response(ok=False, error=str(e))


# ─────────────────────────────────────────
# API: Export Excel (nhận data từ frontend)
# ─────────────────────────────────────────
@bp.route('/api/export_excel', methods=['POST'])
def api_export_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        body = request.get_json(force=True)
        rows = body.get('rows', [])
        col_headers = body.get('col_headers', [])
        title = body.get('title', 'Bán Ra')

        if not rows:
            return api_response(ok=False, error='Không có dữ liệu', status_code=400)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Bán Ra'

        # Header style
        hfont = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        hfill = PatternFill(start_color='1A46C4', end_color='1A46C4', fill_type='solid')
        halign = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin = Side(style='thin', color='D5D9E4')
        hborder = Border(bottom=thin, right=thin)

        # Write headers
        for ci, h in enumerate(col_headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = hfont
            cell.fill = hfill
            cell.alignment = halign
            cell.border = hborder

        # Data style
        dfont = Font(name='Arial', size=10)
        nfont = Font(name='Consolas', size=10)

        # Write data
        for ri, row in enumerate(rows, 2):
            for ci, val in enumerate(row, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font = dfont
                if isinstance(val, (int, float)):
                    cell.font = nfont
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal='right')

        # Auto-width
        for ci in range(1, len(col_headers) + 1):
            max_len = len(str(col_headers[ci - 1]))
            for ri in range(2, min(len(rows) + 2, 102)):
                val = ws.cell(row=ri, column=ci).value
                if val:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 3, 40)

        # Freeze header
        ws.freeze_panes = 'A2'

        # Format as Table
        from openpyxl.worksheet.table import Table, TableStyleInfo
        if col_headers and rows:
            last_col = get_column_letter(len(col_headers))
            last_row = len(rows) + 1

            seen = {}
            for ci, h in enumerate(col_headers, 1):
                safe = str(h).replace('\n', ' ').strip()
                if safe in seen:
                    seen[safe] += 1
                    safe = f'{safe}_{seen[safe]}'
                else:
                    seen[safe] = 0
                ws.cell(row=1, column=ci).value = safe

            tab = Table(displayName='BanRa', ref=f'A1:{last_col}{last_row}')
            tab.tableStyleInfo = TableStyleInfo(
                name='TableStyleMedium2', showFirstColumn=False,
                showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            ws.add_table(tab)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        fn = f'BanRa_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'

        set_api_result(
            status='ok',
            row_count=len(rows),
            meta={'export': fn, 'title': title}
        )

        return send_file(buf, as_attachment=True, download_name=fn,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        logger.exception('export error')
        return api_response(ok=False, error=str(e))