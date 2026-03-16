"""
VietAnh BI — Reports module
Chạy truy vấn SQL trực tiếp, trực quan hóa, xuất Excel

Cách thêm báo cáo mới:
1. Tạo file .sql trong reports/queries/
2. Thêm entry vào REPORT_REGISTRY bên dưới
3. Phân quyền qua trang admin (tương tự dashboard)
"""
from flask import Blueprint, render_template, request, jsonify, abort, g, send_file
from db import get_db
from auth import login_required
from config import DB_TYPE
import os
import io

bp = Blueprint('reports', __name__, url_prefix='/reports')

# Thư mục chứa file SQL
QUERIES_DIR = os.path.join(os.path.dirname(__file__), 'queries')


# ==============================================================
# ĐĂNG KÝ BÁO CÁO — thêm báo cáo mới ở đây
# ==============================================================
# Mỗi báo cáo gồm:
#   - slug: URL-friendly ID
#   - name: tên hiển thị
#   - description: mô tả ngắn
#   - sql_file: tên file trong queries/
#   - params: danh sách tham số [{name, label, type, default}]
#
# VD khi thêm báo cáo mới:
# {
#     'slug': 'doanh-thu-theo-thang',
#     'name': 'Doanh thu theo tháng',
#     'description': 'Tổng hợp doanh thu theo tháng, theo NVKD',
#     'sql_file': 'doanh_thu_theo_thang.sql',
#     'params': [
#         {'name': 'from_date', 'label': 'Từ ngày', 'type': 'date', 'default': '2025-01-01'},
#         {'name': 'to_date', 'label': 'Đến ngày', 'type': 'date', 'default': '2025-12-31'},
#     ]
# },

REPORT_REGISTRY = [
    # Thêm báo cáo ở đây, VD:
    # {
    #     'slug': 'demo-report',
    #     'name': 'Báo cáo Demo',
    #     'description': 'Báo cáo thử nghiệm',
    #     'sql_file': 'demo.sql',
    #     'params': []
    # },
]


def _get_report(slug):
    for r in REPORT_REGISTRY:
        if r['slug'] == slug:
            return r
    return None


def _read_sql(filename):
    path = os.path.join(QUERIES_DIR, filename)
    if not os.path.exists(path):
        return None
    with open(path, 'r', encoding='utf-8') as f:
        return f.read()


@bp.route('/')
@login_required
def report_list():
    return render_template('reports/report_list.html',
        reports=REPORT_REGISTRY,
        username=g.current_user['display_name'] or g.current_user['username'],
        role=g.current_user['role'])


@bp.route('/<slug>')
@login_required
def report_view(slug):
    report = _get_report(slug)
    if not report:
        abort(404)
    return render_template('reports/report_view.html',
        report=report,
        username=g.current_user['display_name'] or g.current_user['username'],
        role=g.current_user['role'])


@bp.route('/api/<slug>/data')
@login_required
def report_data(slug):
    """Chạy truy vấn SQL và trả về JSON"""
    report = _get_report(slug)
    if not report:
        abort(404)

    sql = _read_sql(report['sql_file'])
    if not sql:
        return jsonify({'error': f'Không tìm thấy file SQL: {report["sql_file"]}'}), 404

    # Thu thập tham số từ query string
    params = []
    for p in report.get('params', []):
        val = request.args.get(p['name'], p.get('default', ''))
        params.append(val)

    try:
        db = get_db()
        result = db.execute(sql, params)
        rows = result.fetchall()

        if not rows:
            return jsonify({'columns': [], 'rows': []})

        columns = list(rows[0].keys()) if hasattr(rows[0], 'keys') else []
        data = []
        for row in rows:
            data.append({col: row[col] for col in columns})

        return jsonify({'columns': columns, 'rows': data})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/api/<slug>/excel')
@login_required
def report_excel(slug):
    """Xuất kết quả báo cáo ra Excel"""
    report = _get_report(slug)
    if not report:
        abort(404)

    sql = _read_sql(report['sql_file'])
    if not sql:
        abort(404)

    params = []
    for p in report.get('params', []):
        params.append(request.args.get(p['name'], p.get('default', '')))

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        db = get_db()
        rows = db.execute(sql, params).fetchall()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = report['name'][:31]

        if rows:
            columns = list(rows[0].keys())

            # Header
            header_font = Font(bold=True, color='FFFFFF', size=11)
            header_fill = PatternFill(start_color='1A1A2E', end_color='1A1A2E', fill_type='solid')
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))

            for col_idx, col_name in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

            # Data
            for row_idx, row in enumerate(rows, 2):
                for col_idx, col_name in enumerate(columns, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=row[col_name])
                    cell.border = thin_border

            # Auto-width
            for col_idx, col_name in enumerate(columns, 1):
                max_len = len(str(col_name))
                for row in rows[:100]:
                    val_len = len(str(row[col_name] or ''))
                    if val_len > max_len:
                        max_len = val_len
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 50)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"{report['slug']}.xlsx"
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=filename)
    except ImportError:
        return jsonify({'error': 'Cần cài openpyxl: pip install openpyxl'}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500