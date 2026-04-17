"""
Báo cáo Chi Tiết Bán Hàng & Thanh Toán — Blueprint (DuckDB version)
API prefix: /reports/bao-cao-chi-tiet/api/...
"""
from flask import Blueprint, request, send_file, current_app
from api_logger import api_response, set_api_result
from datetime import datetime, date
import logging

logger = logging.getLogger(__name__)

bp = Blueprint('bcct', __name__,
               url_prefix='/reports/bao-cao-chi-tiet',
               template_folder='templates')
bp.api_report = 'bao-cao-chi-tiet'
from query_loader import load_sql


def get_store():
    return current_app.config['DUCKDB_STORE']


# ─────────────────────────────────────────
# API: Hierarchy
# ─────────────────────────────────────────
@bp.route('/api/hierarchy')
def api_hierarchy():
    try:
        data = get_store().query(load_sql('HIERARCHY_CTE_DUCK'))
        return api_response(ok=True, data=data, count=len(data))
    except Exception as e:
        logger.error(f"[BCCT hierarchy] {e}")
        return api_response(ok=False, error=str(e))


# ─────────────────────────────────────────
# API: Khách hàng
# ─────────────────────────────────────────
@bp.route('/api/khachhang')
def api_khachhang():
    try:
        data = get_store().query(load_sql('KHACHHANG_DUCK'))
        return api_response(ok=True, data=data, count=len(data))
    except Exception as e:
        logger.error(f"[BCCT khachhang] {e}")
        return api_response(ok=False, error=str(e))


# ─────────────────────────────────────────
# API: Doanh số chi tiết
# ─────────────────────────────────────────
@bp.route('/api/doanhso_chitiet', methods=['POST'])
def api_doanhso_chitiet():
    body = request.get_json(force=True)
    ngay_a = body.get('ngay_a', '')
    ngay_b = body.get('ngay_b', '')
    ma_bp = body.get('ma_bp', '')
    ds_nvkd = body.get('ds_nvkd', '')
    ds_kh = body.get('ds_kh', '')

    if not ngay_a or not ngay_b:
        return api_response(ok=False, error='Thiếu ngày', status_code=400)

    try:
        data = get_store().query(
            load_sql('DOANHSO_CHITIET_BCCT_DUCK'),
            [ngay_a, ngay_b, ma_bp or '', ds_nvkd or '', ds_kh or '']
        )
        for d in data:
            for k in ('so_luong', 'gia_nt2', 'tien_nt2', 'tien_ck_nt', 'ts_gtgt', 'thue_gtgt_nt', 'doanhso'):
                if d.get(k) is not None:
                    d[k] = float(d[k])
        return api_response(ok=True, data=data, count=len(data),
                            meta={'ngay_a': ngay_a, 'ngay_b': ngay_b, 'ma_bp': ma_bp})
    except Exception as e:
        logger.error(f"[BCCT doanhso] {e}")
        return api_response(ok=False, error=str(e))


# ─────────────────────────────────────────
# API: Doanh thu chi tiết
# ─────────────────────────────────────────
@bp.route('/api/doanhthu_chitiet', methods=['POST'])
def api_doanhthu_chitiet():
    body = request.get_json(force=True)
    ngay_a = body.get('ngay_a', '')
    ngay_b = body.get('ngay_b', '')
    ma_bp = body.get('ma_bp', '')
    ds_nvkd = body.get('ds_nvkd', '')
    ds_kh = body.get('ds_kh', '')

    if not ngay_a or not ngay_b:
        return api_response(ok=False, error='Thiếu ngày', status_code=400)

    try:
        data = get_store().query(
            load_sql('DOANHTHU_CHITIET_BCCT_DUCK'),
            [ngay_a, ngay_b, ma_bp or '', ds_nvkd or '', ds_kh or '']
        )
        for d in data:
            if d.get('doanhthu') is not None:
                d['doanhthu'] = float(d['doanhthu'])
        return api_response(ok=True, data=data, count=len(data),
                            meta={'ngay_a': ngay_a, 'ngay_b': ngay_b, 'ma_bp': ma_bp})
    except Exception as e:
        logger.error(f"[BCCT doanhthu] {e}")
        return api_response(ok=False, error=str(e))


# ─────────────────────────────────────────
# API: Export Excel
# ─────────────────────────────────────────
@bp.route('/api/export_excel', methods=['POST'])
def api_export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    body = request.get_json(force=True)
    rows = body.get('rows', [])
    col_headers = body.get('col_headers', [])

    if not rows:
        return api_response(ok=False, error='Không có dữ liệu', status_code=400)

    max_nv_depth = max((r.get('depth', 0) for r in rows if r.get('type') == 'nv'), default=0)
    nv_cols = max_nv_depth + 1
    kh_col = nv_cols + 1
    data_start = nv_cols + 2

    wb = Workbook()
    ws = wb.active
    ws.title = 'Chi tiết'
    FN = 'Arial'
    BK = '000000'
    thin = Border(bottom=Side(style='thin', color='D5D9E4'), right=Side(style='thin', color='ECEEF3'))
    hdr = Border(bottom=Side(style='medium', color='A0AAC0'))
    tot = Border(top=Side(style='medium', color='8090B0'), bottom=Side(style='medium', color='8090B0'))
    bgs = ['B8C6F0', 'CADAF6', 'DAEAFC', 'E8F0FD', 'F0F5FE', 'F7FAFE']
    kh_bg = 'FFF8E1'
    ds_bg = 'FFFFFF'
    dt_bg = 'F0FFF0'

    def nbg(d): return bgs[min(d, len(bgs) - 1)]
    def nsz(d): return 11 if d == 0 else 10.5 if d == 1 else 10

    cur_row = 1
    for c in range(1, nv_cols + 1):
        cell = ws.cell(row=1, column=c, value='NHÂN VIÊN KD' if c == 1 else '')
        cell.font = Font(name=FN, bold=True, size=11, color=BK)
        cell.fill = PatternFill('solid', fgColor='D0D8ED')
        cell.alignment = Alignment(vertical='center')
        cell.border = hdr
    kh = ws.cell(row=1, column=kh_col, value='KHÁCH HÀNG')
    kh.font = Font(name=FN, bold=True, size=11, color=BK)
    kh.fill = PatternFill('solid', fgColor='D0D8ED')
    kh.alignment = Alignment(vertical='center')
    kh.border = hdr
    for ci, ch in enumerate(col_headers):
        cell = ws.cell(row=1, column=data_start + ci, value=ch)
        cell.font = Font(name=FN, bold=True, size=10, color=BK)
        cell.fill = PatternFill('solid', fgColor='D0D8ED')
        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        cell.border = hdr
    ws.row_dimensions[1].height = 30

    for rd in rows:
        cur_row += 1
        rt = rd.get('type', '')
        dp = rd.get('depth', 0)
        nm = rd.get('name', '')
        vs = rd.get('values', [])
        tc = data_start + len(col_headers) - 1

        if rt == 'nv':
            ws.cell(row=cur_row, column=min(dp, nv_cols - 1) + 1, value=nm)
            bg = nbg(dp)
            sz = nsz(dp)
            for c in range(1, tc + 1):
                cell = ws.cell(row=cur_row, column=c)
                cell.fill = PatternFill('solid', fgColor=bg)
                cell.border = thin
                cell.font = Font(name=FN, bold=True, size=sz, color=BK)
                cell.alignment = Alignment(vertical='center', horizontal='right' if c >= data_start else 'left')
                if c >= data_start:
                    cell.number_format = '#,##0'
            for vi, v in enumerate(vs):
                if v is not None and v != '':
                    ws.cell(row=cur_row, column=data_start + vi, value=v)

        elif rt == 'kh':
            ws.cell(row=cur_row, column=kh_col, value=nm)
            for c in range(1, tc + 1):
                cell = ws.cell(row=cur_row, column=c)
                cell.fill = PatternFill('solid', fgColor=kh_bg)
                cell.border = thin
                cell.font = Font(name=FN, bold=True, size=10, color='6B5B00')
                cell.alignment = Alignment(vertical='center', horizontal='right' if c >= data_start else 'left')
                if c >= data_start:
                    cell.number_format = '#,##0'
            for vi, v in enumerate(vs):
                if v is not None and v != '':
                    ws.cell(row=cur_row, column=data_start + vi, value=v)

        elif rt in ('ds', 'dt'):
            bg_color = ds_bg if rt == 'ds' else dt_bg
            for c in range(1, tc + 1):
                cell = ws.cell(row=cur_row, column=c)
                cell.fill = PatternFill('solid', fgColor=bg_color)
                cell.border = thin
                cell.font = Font(name=FN, size=9.5, color=BK)
                cell.alignment = Alignment(vertical='center', horizontal='right' if c >= data_start else 'left')
                if c >= data_start:
                    cell.number_format = '#,##0'
            for vi, v in enumerate(vs):
                if v is not None and v != '':
                    ws.cell(row=cur_row, column=data_start + vi, value=v)

        elif rt == 'total':
            ws.cell(row=cur_row, column=1, value='TỔNG CỘNG')
            for c in range(1, tc + 1):
                cell = ws.cell(row=cur_row, column=c)
                cell.fill = PatternFill('solid', fgColor='B8C6F0')
                cell.border = tot
                cell.font = Font(name=FN, bold=True, size=11, color=BK)
                cell.alignment = Alignment(vertical='center', horizontal='right' if c >= data_start else 'left')
                if c >= data_start:
                    cell.number_format = '#,##0'
            for vi, v in enumerate(vs):
                if v is not None and v != '':
                    ws.cell(row=cur_row, column=data_start + vi, value=v)

        ws.row_dimensions[cur_row].height = 18

    for c in range(1, nv_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 5
    ws.column_dimensions[get_column_letter(kh_col)].width = 28
    for ci in range(len(col_headers)):
        ws.column_dimensions[get_column_letter(data_start + ci)].width = 16
    ws.freeze_panes = ws.cell(row=2, column=data_start)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    now = datetime.now()
    filename = f'BaoCaoChiTiet_{now.strftime("%Y%m%d")}.xlsx'

    set_api_result(
        status='ok',
        row_count=len(rows),
        meta={'export': filename}
    )

    return send_file(buf,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)