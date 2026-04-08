"""admin/kpi.py — KPI management: phân công + KPI targets (SQL Server)"""
import json
from flask import request, jsonify
from database import get_db
from auth import admin_required
from admin import bp

try:
    from config import SQLSERVER_CONFIG, DB_TYPE
except ImportError:
    SQLSERVER_CONFIG = None
    DB_TYPE = 'sqlite'


def _ss_conn():
    """Get direct SQL Server connection for kpi_targets (lives on SQL Server)."""
    import pyodbc
    c = SQLSERVER_CONFIG
    return pyodbc.connect(
        f"DRIVER={{{c['driver']}}};SERVER={c['server']},{c['port']};"
        f"DATABASE={c['database']};UID={c['username']};PWD={c['password']};"
        "TrustServerCertificate=yes;Connect Timeout=30;",
        autocommit=False
    )


def _recalc_paths(conn, ma_kbc):
    """Recalc stt_nhom for all rows in a period."""
    cur = conn.cursor()
    cur.execute('SELECT ma_nvkd, ma_ql, ma_bp FROM kpi_targets WHERE ma_kbc = ?', (ma_kbc,))
    rows = cur.fetchall()
    by_key = {}
    for r in rows:
        by_key[r[0]] = {'ma_ql': r[1] or '', 'ma_bp': r[2] or ''}

    for ma in by_key:
        path = []
        cur_ma = ma
        visited = set()
        while cur_ma and cur_ma in by_key and cur_ma not in visited:
            path.append(cur_ma)
            visited.add(cur_ma)
            cur_ma = by_key[cur_ma]['ma_ql']
        path.reverse()
        bp_code = by_key[ma]['ma_bp']
        stt = 'VAG.' + bp_code + '00.' + '.'.join(path) if path else ''
        cur.execute('UPDATE kpi_targets SET stt_nhom = ?, ldate = GETDATE() WHERE ma_kbc = ? AND ma_nvkd = ?',
                    (stt, ma_kbc, ma))
    conn.commit()


def _parse_kbc(ma_kbc):
    """Parse T01-2026 → (2026, 1)"""
    try:
        parts = ma_kbc.replace('T', '').split('-')
        return int(parts[1]), int(parts[0])
    except:
        return 2026, 1


# ═══════════════════════════════════════════════
# API: Load data
# ═══════════════════════════════════════════════

@bp.route('/kpi/data')
@admin_required
def kpi_data():
    """Load KPI + unassigned NVKDs in 1 request."""
    ma_kbc_list = request.args.getlist('ma_kbc')
    ma_bp = request.args.get('ma_bp', '').strip()

    kbcs = []
    for k in ma_kbc_list:
        for part in k.split(','):
            part = part.strip()
            if part:
                kbcs.append(part)

    if not kbcs:
        return jsonify({'ok': True, 'assignments': [], 'kpi': {}, 'detail': {}, 'unassigned': []})

    try:
        conn = _ss_conn()
        cur = conn.cursor()

        # 1. Load KPI data
        ph = ','.join(['?'] * len(kbcs))
        sql = f'''SELECT ma_kbc, ma_nvkd, ten_nvkd, ma_ql, ma_bp, stt_nhom,
                         kpi, kpi_cong_ty, kpi_ds, kpi_ds_cong_ty
                  FROM kpi_targets WHERE ma_kbc IN ({ph})'''
        params = list(kbcs)
        if ma_bp:
            sql += ' AND ma_bp = ?'
            params.append(ma_bp)
        sql += ' ORDER BY stt_nhom, ma_nvkd'
        cur.execute(sql, params)
        cols = [d[0] for d in cur.description]
        rows = cur.fetchall()

        seen_nv = {}
        kpi_sum = {}
        detail = {}
        flds = ['kpi', 'kpi_cong_ty', 'kpi_ds', 'kpi_ds_cong_ty']

        for r in rows:
            rd = {cols[i]: r[i] for i in range(len(cols))}
            ma = rd['ma_nvkd'] or ''
            if not ma:
                continue
            kbc = rd['ma_kbc'] or ''
            if ma not in seen_nv:
                seen_nv[ma] = {
                    'ma_nvkd': ma, 'ten_nvkd': rd['ten_nvkd'] or '',
                    'ma_ql': rd['ma_ql'] or '', 'ma_bp': rd['ma_bp'] or '',
                    'stt_nhom': rd['stt_nhom'] or '',
                }
            if ma not in kpi_sum:
                kpi_sum[ma] = {f: 0 for f in flds}
            for f in flds:
                kpi_sum[ma][f] += float(rd[f] or 0)
            if ma not in detail:
                detail[ma] = {}
            detail[ma][kbc] = {f: float(rd[f] or 0) for f in flds}

        existing_nvkds = set(seen_nv.keys())

        # 2. Load unassigned NVKDs (all from DMNHANVIENKD_VIEW not in kpi_targets)
        cur.execute("""SELECT ma_nvkd, ten_nvkd, ma_ql
                       FROM DMNHANVIENKD_VIEW
                       WHERE ma_nvkd IS NOT NULL AND ma_nvkd != ''""")
        all_nv = cur.fetchall()

        bp_map = {}
        cur.execute("""SELECT DISTINCT ma_nvkd, ma_bp FROM DMKHACHHANG_VIEW
                       WHERE ma_nvkd IS NOT NULL AND ma_nvkd != ''
                         AND ma_bp IS NOT NULL AND ma_bp != '' AND ma_bp != 'TN'""")
        for r in cur.fetchall():
            bp_map[r[0]] = r[1]

        conn.close()

        unassigned = []
        for r in all_nv:
            nv_ma = (r[0] or '').strip()
            if not nv_ma or nv_ma in existing_nvkds:
                continue
            nv_bp = bp_map.get(nv_ma, '')
            if ma_bp and nv_bp != ma_bp:
                continue
            unassigned.append({
                'ma_nvkd': nv_ma, 'ten_nvkd': (r[1] or '').strip(),
                'ma_ql': (r[2] or '').strip(), 'ma_bp': nv_bp
            })

        # 3. Load ratios
        ratios = {}
        try:
            cur2 = conn.cursor() if not conn.closed else _ss_conn().cursor()
            ph2 = ','.join(['?'] * len(kbcs))
            sql_r = f'SELECT ma_bp, ma_kbc, ratio_dt_cty, ratio_ds_nb, ratio_ds_cty FROM kpi_ratios WHERE ma_kbc IN ({ph2})'
            params_r = list(kbcs)
            if ma_bp:
                sql_r += ' AND ma_bp = ?'
                params_r.append(ma_bp)
            cur2.execute(sql_r, params_r)
            for r in cur2.fetchall():
                bp_c = r[0] or ''
                kbc_c = r[1] or ''
                if bp_c not in ratios:
                    ratios[bp_c] = {}
                ratios[bp_c][kbc_c] = {
                    'ratio_dt_cty': float(r[2] or 1),
                    'ratio_ds_nb': float(r[3] or 1),
                    'ratio_ds_cty': float(r[4] or 1),
                }
        except:
            pass

        # 4. Auto-calc missing ratios from actual data
        # For each BP+kbc: if ratios not in table but NVs have kpi>0 and kpi_cong_ty>0, calc ratio
        try:
            bp_kbc_kpi = {}  # {(bp, kbc): {sum_kpi, sum_cty, sum_ds, sum_ds_cty}}
            for ma_nv in detail:
                nv_info = seen_nv.get(ma_nv)
                if not nv_info:
                    continue
                nv_bp = nv_info.get('ma_bp', '')
                if not nv_bp:
                    continue
                for kbc_k, kpi_vals in detail[ma_nv].items():
                    key = (nv_bp, kbc_k)
                    if key not in bp_kbc_kpi:
                        bp_kbc_kpi[key] = {'kpi': 0, 'cty': 0, 'ds': 0, 'ds_cty': 0}
                    bp_kbc_kpi[key]['kpi'] += kpi_vals.get('kpi', 0)
                    bp_kbc_kpi[key]['cty'] += kpi_vals.get('kpi_cong_ty', 0)
                    bp_kbc_kpi[key]['ds'] += kpi_vals.get('kpi_ds', 0)
                    bp_kbc_kpi[key]['ds_cty'] += kpi_vals.get('kpi_ds_cong_ty', 0)

            new_ratios = []
            for (bp_c, kbc_c), sums in bp_kbc_kpi.items():
                # Skip if already have ratio for this BP+kbc
                if bp_c in ratios and kbc_c in ratios[bp_c]:
                    continue
                if sums['kpi'] <= 0:
                    continue
                r_dt_cty = sums['cty'] / sums['kpi'] if sums['cty'] > 0 else 1
                r_ds_nb = sums['ds'] / sums['kpi'] if sums['ds'] > 0 else 1
                r_ds_cty = sums['ds_cty'] / sums['kpi'] if sums['ds_cty'] > 0 else 1
                # Save to ratios dict
                if bp_c not in ratios:
                    ratios[bp_c] = {}
                ratios[bp_c][kbc_c] = {
                    'ratio_dt_cty': round(r_dt_cty, 6),
                    'ratio_ds_nb': round(r_ds_nb, 6),
                    'ratio_ds_cty': round(r_ds_cty, 6),
                }
                new_ratios.append((bp_c, kbc_c, r_dt_cty, r_ds_nb, r_ds_cty))

            # Upsert new ratios into DB
            if new_ratios:
                try:
                    cur3 = conn.cursor() if not conn.closed else _ss_conn().cursor()
                    for bp_c, kbc_c, r1, r2, r3 in new_ratios:
                        nam_r, thang_r = _parse_kbc(kbc_c)
                        cur3.execute('SELECT id FROM kpi_ratios WHERE ma_bp=? AND ma_kbc=?', (bp_c, kbc_c))
                        if not cur3.fetchone():
                            cur3.execute('''INSERT INTO kpi_ratios (ma_bp, ma_kbc, nam, thang, ratio_dt_cty, ratio_ds_nb, ratio_ds_cty)
                                            VALUES (?, ?, ?, ?, ?, ?, ?)''',
                                         (bp_c, kbc_c, nam_r, thang_r, r1, r2, r3))
                    conn.commit()
                except:
                    pass
        except:
            pass

        if not conn.closed:
            conn.close()

        return jsonify({
            'ok': True, 'assignments': list(seen_nv.values()),
            'kpi': kpi_sum, 'detail': detail, 'unassigned': unassigned,
            'ratios': ratios
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


# ═══════════════════════════════════════════════
# API: Save single cell
# ═══════════════════════════════════════════════

@bp.route('/kpi/save-cell', methods=['POST'])
@admin_required
def kpi_save_cell():
    data = request.get_json(force=True)
    ma_kbc = data.get('ma_kbc', '').strip()
    ma_nvkd = data.get('ma_nvkd', '').strip()
    field = data.get('field', '')
    value = data.get('value', 0)

    if not ma_nvkd:
        return jsonify({'ok': False, 'error': 'Missing ma_nvkd'}), 400
    if field not in ('kpi', 'kpi_cong_ty', 'kpi_ds', 'kpi_ds_cong_ty'):
        return jsonify({'ok': False, 'error': 'Invalid field'}), 400

    try:
        conn = _ss_conn()
        cur = conn.cursor()
        cur.execute('SELECT id, ma_bp FROM kpi_targets WHERE ma_kbc = ? AND ma_nvkd = ?', (ma_kbc, ma_nvkd))
        row = cur.fetchone()

        if row:
            cur.execute(f'UPDATE kpi_targets SET {field} = ?, ldate = GETDATE() WHERE ma_kbc = ? AND ma_nvkd = ?',
                        (value, ma_kbc, ma_nvkd))
            nv_bp = row[1] or ''
        else:
            nam, thang = _parse_kbc(ma_kbc)
            cur.execute(
                f'INSERT INTO kpi_targets (ma_bp, ma_kbc, nam, thang, ma_nvkd, {field}) VALUES (?, ?, ?, ?, ?, ?)',
                ('', ma_kbc, nam, thang, ma_nvkd, value))
            nv_bp = ''

        # If field is 'kpi' (DT nội bộ), auto-calc other 3 fields using ratios
        calc = {}
        if field == 'kpi' and nv_bp:
            cur.execute('SELECT ratio_dt_cty, ratio_ds_nb, ratio_ds_cty FROM kpi_ratios WHERE ma_bp = ? AND ma_kbc = ?',
                        (nv_bp, ma_kbc))
            rr = cur.fetchone()
            if rr:
                calc = {
                    'kpi_cong_ty': round(value * float(rr[0] or 1), 2),
                    'kpi_ds': round(value * float(rr[1] or 1), 2),
                    'kpi_ds_cong_ty': round(value * float(rr[2] or 1), 2),
                }
                cur.execute('''UPDATE kpi_targets SET kpi_cong_ty=?, kpi_ds=?, kpi_ds_cong_ty=?, ldate=GETDATE()
                               WHERE ma_kbc=? AND ma_nvkd=?''',
                            (calc['kpi_cong_ty'], calc['kpi_ds'], calc['kpi_ds_cong_ty'], ma_kbc, ma_nvkd))

        conn.commit()
        conn.close()
        return jsonify({'ok': True, 'calc': calc})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


# ═══════════════════════════════════════════════
# API: Reassign (drag & drop)
# ═══════════════════════════════════════════════

@bp.route('/kpi/reassign', methods=['POST'])
@admin_required
def kpi_reassign():
    data = request.get_json(force=True)
    ma_kbc = data.get('ma_kbc', '').strip()
    ma_nvkd = data.get('ma_nvkd', '').strip()
    new_ma_ql = data.get('new_ma_ql', '').strip()

    if not ma_nvkd:
        return jsonify({'ok': False}), 400

    try:
        conn = _ss_conn()
        cur = conn.cursor()
        cur.execute('UPDATE kpi_targets SET ma_ql = ?, ldate = GETDATE() WHERE ma_kbc = ? AND ma_nvkd = ?',
                    (new_ma_ql, ma_kbc, ma_nvkd))
        conn.commit()
        # Recalc paths
        _recalc_paths(conn, ma_kbc)

        # Return updated assignments
        ma_bp = data.get('ma_bp', '')
        sql = 'SELECT ma_nvkd, ten_nvkd, ma_ql, ma_bp, stt_nhom FROM kpi_targets WHERE ma_kbc = ?'
        params = [ma_kbc]
        if ma_bp:
            sql += ' AND ma_bp = ?'
            params.append(ma_bp)
        cur.execute(sql, params)
        cols = [d[0] for d in cur.description]
        rows = cur.fetchall()
        conn.close()

        assignments = []
        for r in rows:
            rd = {cols[i]: r[i] for i in range(len(cols))}
            assignments.append({
                'ma_nvkd': rd['ma_nvkd'] or '',
                'ten_nvkd': rd['ten_nvkd'] or '',
                'ma_ql': rd['ma_ql'] or '',
                'ma_bp': rd['ma_bp'] or '',
                'stt_nhom': rd['stt_nhom'] or '',
            })
        return jsonify({'ok': True, 'assignments': assignments})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


# ═══════════════════════════════════════════════
# API: Copy period (clone kỳ trước → kỳ mới)
# ═══════════════════════════════════════════════

@bp.route('/kpi/copy-period', methods=['POST'])
@admin_required
def kpi_copy_period():
    """Copy kiến trúc (phân công) từ kỳ nguồn sang kỳ đích.
    - Chỉ copy cấu trúc (ma_ql, stt_nhom), KHÔNG copy số KPI (set = 0)
    - NV đã tồn tại trong kỳ đích → chỉ update ma_ql + stt_nhom, giữ nguyên KPI
    - NV chưa có → insert mới với KPI = 0
    """
    data = request.get_json(force=True)
    from_kbc = data.get('from_kbc', '').strip()
    to_kbc = data.get('to_kbc', '').strip()
    ma_bp = data.get('ma_bp', '').strip()

    if not from_kbc or not to_kbc:
        return jsonify({'ok': False, 'error': 'Thiếu from_kbc hoặc to_kbc'}), 400
    if from_kbc == to_kbc:
        return jsonify({'ok': False, 'error': 'Kỳ nguồn và đích không được trùng'}), 400

    nam, thang = _parse_kbc(to_kbc)

    try:
        conn = _ss_conn()
        cur = conn.cursor()

        # Load source assignments
        sql_src = '''SELECT ma_nvkd, ten_nvkd, nguoi_gd, ma_ql, stt_nhom, ma_bp
                     FROM kpi_targets WHERE ma_kbc = ?'''
        params_src = [from_kbc]
        if ma_bp:
            sql_src += ' AND ma_bp = ?'
            params_src.append(ma_bp)
        cur.execute(sql_src, params_src)
        src_cols = [d[0] for d in cur.description]
        src_rows = cur.fetchall()

        # Load existing NVKDs in target
        cur.execute('SELECT ma_nvkd FROM kpi_targets WHERE ma_kbc = ?', (to_kbc,))
        existing = set(r[0] for r in cur.fetchall())

        inserted = 0
        updated = 0
        for r in src_rows:
            rd = {src_cols[i]: r[i] for i in range(len(src_cols))}
            ma = rd['ma_nvkd'] or ''
            if not ma:
                continue

            if ma in existing:
                # Update structure only (keep KPI numbers)
                cur.execute('''UPDATE kpi_targets SET ma_ql = ?, stt_nhom = ?, ldate = GETDATE()
                               WHERE ma_kbc = ? AND ma_nvkd = ?''',
                            (rd['ma_ql'] or '', rd['stt_nhom'] or '', to_kbc, ma))
                updated += 1
            else:
                # Insert new with KPI = 0
                cur.execute('''INSERT INTO kpi_targets
                    (ma_bp, ma_kbc, nam, thang, ma_nvkd, ten_nvkd, nguoi_gd, ma_ql, stt_nhom,
                     kpi, kpi_cong_ty, kpi_ds, kpi_ds_cong_ty, merge_kpi)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 0, 0, 0, 0, 0)''',
                    (rd['ma_bp'] or ma_bp, to_kbc, nam, thang, ma,
                     rd['ten_nvkd'] or '', rd['nguoi_gd'] or '',
                     rd['ma_ql'] or '', rd['stt_nhom'] or ''))
                inserted += 1

        conn.commit()
        _recalc_paths(conn, to_kbc)
        conn.close()
        return jsonify({
            'ok': True,
            'message': f'Copy kiến trúc {from_kbc} → {to_kbc}: +{inserted} mới, ~{updated} cập nhật cấu trúc (KPI giữ nguyên)'
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


# ═══════════════════════════════════════════════
# API: Import from DMNHANVIENKD_VIEW (hệ thống)
# ═══════════════════════════════════════════════

@bp.route('/kpi/import-system', methods=['POST'])
@admin_required
def kpi_import_system():
    data = request.get_json(force=True)
    ma_kbc = data.get('ma_kbc', '').strip()
    ma_bp = data.get('ma_bp', '').strip()

    if not ma_kbc:
        return jsonify({'ok': False, 'error': 'Thiếu ma_kbc'}), 400

    nam, thang = _parse_kbc(ma_kbc)

    try:
        conn = _ss_conn()
        cur = conn.cursor()

        # Get existing NVKDs for this period (to skip)
        cur.execute('SELECT ma_nvkd FROM kpi_targets WHERE ma_kbc = ?', (ma_kbc,))
        existing = set(r[0] for r in cur.fetchall())

        # Query current employees from system view
        sql_nv = "SELECT ma_nvkd, ten_nvkd, ma_ql FROM DMNHANVIENKD_VIEW WHERE ma_nvkd IS NOT NULL AND ma_nvkd != ''"
        cur.execute(sql_nv)
        nv_rows = cur.fetchall()

        # Determine BP for each NV from DMKHACHHANG_VIEW
        bp_map = {}
        if ma_bp:
            bp_map = {r[0]: ma_bp for r in nv_rows}  # all same BP
        else:
            cur.execute("""SELECT DISTINCT ma_nvkd, ma_bp FROM DMKHACHHANG_VIEW
                           WHERE ma_nvkd IS NOT NULL AND ma_nvkd != ''
                             AND ma_bp IS NOT NULL AND ma_bp != '' AND ma_bp != 'TN'""")
            for r in cur.fetchall():
                bp_map[r[0]] = r[1]

        inserted = 0
        skipped = 0
        for r in nv_rows:
            nv_ma = (r[0] or '').strip()
            if not nv_ma or nv_ma in existing:
                skipped += 1
                continue
            nv_ten = (r[1] or '').strip()
            nv_ql = (r[2] or '').strip()
            nv_bp = bp_map.get(nv_ma, ma_bp or '')
            # Fallback: no BP → XX
            if not nv_bp:
                nv_bp = 'XX'
            # Fallback: has BP but no ma_ql → ma_ql = BP + "99"
            if nv_bp and not nv_ql:
                nv_ql = nv_bp + '99'

            if ma_bp and nv_bp != ma_bp:
                continue  # filter by BP

            cur.execute(
                '''INSERT INTO kpi_targets (ma_bp, ma_kbc, nam, thang, ma_nvkd, ten_nvkd, nguoi_gd, ma_ql)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
                (nv_bp, ma_kbc, nam, thang, nv_ma, nv_ten, nv_ten, nv_ql))
            inserted += 1

        conn.commit()
        # Recalc paths
        _recalc_paths(conn, ma_kbc)
        conn.close()
        return jsonify({'ok': True, 'message': f'Đã import {inserted} NV, bỏ qua {skipped} đã tồn tại'})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


# ═══════════════════════════════════════════════
# API: Export Excel
# ═══════════════════════════════════════════════

@bp.route('/kpi/export')
@admin_required
def kpi_export():
    from flask import send_file
    from io import BytesIO
    import openpyxl

    ma_kbc = request.args.get('ma_kbc', '').strip()
    ma_bp = request.args.get('ma_bp', '').strip()

    if not ma_kbc:
        return jsonify({'ok': False, 'error': 'Thiếu ma_kbc'}), 400

    try:
        conn = _ss_conn()
        cur = conn.cursor()
        sql = '''SELECT ma_bp, ma_kbc, nam, thang, ten_nvkd, nguoi_gd, ma_nvkd, ma_ql, stt_nhom,
                        kpi, merge_kpi, kpi_cong_ty, kpi_ds, kpi_ds_cong_ty, cdate
                 FROM kpi_targets WHERE ma_kbc = ?'''
        params = [ma_kbc]
        if ma_bp:
            sql += ' AND ma_bp = ?'
            params.append(ma_bp)
        sql += ' ORDER BY stt_nhom, ma_nvkd'
        cur.execute(sql, params)
        cols = [d[0] for d in cur.description]
        rows = cur.fetchall()
        conn.close()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = ma_kbc
        ws.append(cols)
        for r in rows:
            ws.append(list(r))

        # Auto column width
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f'KPI_{ma_kbc}_{ma_bp or "ALL"}.xlsx'
        return send_file(buf, download_name=fname, as_attachment=True,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


# ═══════════════════════════════════════════════
# API: Import Excel
# ═══════════════════════════════════════════════

@bp.route('/kpi/import-excel', methods=['POST'])
@admin_required
def kpi_import_excel():
    """Import KPI từ file Excel.
    Cột bắt buộc: ma_nvkd, thang.
    ma_kbc tự build từ thang+nam nếu cột là formula.
    Mode: upsert (default) | skip.
    """
    import openpyxl
    from datetime import datetime as dt

    f = request.files.get('file')
    if not f:
        return jsonify({'ok': False, 'error': 'Không có file'}), 400

    mode = request.form.get('mode', 'upsert')

    try:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=False)
        ws = wb[wb.sheetnames[0]]

        header = None
        for row in ws.iter_rows(max_row=1, values_only=True):
            header = [str(c or '').strip().lower() for c in row]
            break
        if not header:
            return jsonify({'ok': False, 'error': 'File rỗng'}), 400

        CI = {}
        for i, h in enumerate(header):
            CI[h] = i

        for req in ['ma_nvkd', 'thang']:
            if req not in CI:
                return jsonify({'ok': False, 'error': f'Thiếu cột: {req}'}), 400

        def g(row, col):
            idx = CI.get(col)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        def sf(v):
            if v is None: return 0
            try: return float(v)
            except: return 0

        def si(v):
            if v is None: return 0
            try: return int(v)
            except: return 0

        def ss(v):
            if v is None: return ''
            s = str(v).strip()
            return '' if s.startswith('=') else s

        records = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            ma_nvkd = ss(g(row, 'ma_nvkd'))
            if not ma_nvkd:
                continue
            thang = si(g(row, 'thang'))
            nam = si(g(row, 'nam')) or 2026
            ma_kbc_raw = ss(g(row, 'ma_kbc'))
            ma_kbc = ma_kbc_raw if ma_kbc_raw else f'T{thang:02d}-{nam}'

            records.append({
                'ma_bp': ss(g(row, 'ma_bp')),
                'ma_kbc': ma_kbc,
                'nam': nam, 'thang': thang,
                'ma_nvkd': ma_nvkd,
                'ten_nvkd': ss(g(row, 'ten_nvkd')),
                'nguoi_gd': ss(g(row, 'nguoi_gd')),
                'ma_ql': ss(g(row, 'ma_ql')),
                'stt_nhom': ss(g(row, 'stt_nhom')),
                'kpi': sf(g(row, 'kpi')),
                'kpi_cong_ty': sf(g(row, 'kpi_cong_ty')),
                'kpi_ds': sf(g(row, 'kpi_ds')),
                'kpi_ds_cong_ty': sf(g(row, 'kpi_ds_cong_ty')),
                'merge_kpi': si(g(row, 'merge_kpi')),
                'cdate': g(row, 'cdate'),
            })

        if not records:
            return jsonify({'ok': False, 'error': 'Không có dữ liệu hợp lệ'}), 400

        conn = _ss_conn()
        cur = conn.cursor()
        ins = upd = skp = 0

        for r in records:
            cur.execute('SELECT id FROM kpi_targets WHERE ma_kbc = ? AND ma_nvkd = ?',
                        (r['ma_kbc'], r['ma_nvkd']))
            exists = cur.fetchone()
            if exists:
                if mode == 'skip':
                    skp += 1
                    continue
                cur.execute('''UPDATE kpi_targets SET
                    ma_bp=?, nam=?, thang=?, ten_nvkd=?, nguoi_gd=?, ma_ql=?, stt_nhom=?,
                    kpi=?, kpi_cong_ty=?, kpi_ds=?, kpi_ds_cong_ty=?, merge_kpi=?, ldate=GETDATE()
                    WHERE ma_kbc=? AND ma_nvkd=?''',
                    (r['ma_bp'], r['nam'], r['thang'], r['ten_nvkd'], r['nguoi_gd'],
                     r['ma_ql'], r['stt_nhom'], r['kpi'], r['kpi_cong_ty'],
                     r['kpi_ds'], r['kpi_ds_cong_ty'], r['merge_kpi'],
                     r['ma_kbc'], r['ma_nvkd']))
                upd += 1
            else:
                cdate = r['cdate'] if isinstance(r['cdate'], dt) else None
                cur.execute('''INSERT INTO kpi_targets
                    (ma_bp, ma_kbc, nam, thang, ma_nvkd, ten_nvkd, nguoi_gd,
                     ma_ql, stt_nhom, kpi, kpi_cong_ty, kpi_ds, kpi_ds_cong_ty,
                     merge_kpi, cdate)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                    (r['ma_bp'], r['ma_kbc'], r['nam'], r['thang'], r['ma_nvkd'],
                     r['ten_nvkd'], r['nguoi_gd'], r['ma_ql'], r['stt_nhom'],
                     r['kpi'], r['kpi_cong_ty'], r['kpi_ds'], r['kpi_ds_cong_ty'],
                     r['merge_kpi'], cdate))
                ins += 1

        conn.commit()
        conn.close()

        kbcs = sorted(set(r['ma_kbc'] for r in records))
        return jsonify({
            'ok': True,
            'message': f'{len(records)} dòng: +{ins} mới, ~{upd} cập nhật, ={skp} bỏ qua. Kỳ: {", ".join(kbcs)}'
        })

    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


# ═══════════════════════════════════════════════
# API: Unassigned NVKDs
# ═══════════════════════════════════════════════

@bp.route('/kpi/unassigned')
@admin_required
def kpi_unassigned():
    """NV active (DMNHANVIENKD_VIEW) nhưng chưa có trong kpi_targets cho kỳ này."""
    ma_kbc_list = request.args.getlist('ma_kbc')
    ma_bp = request.args.get('ma_bp', '').strip()
    kbcs = []
    for k in ma_kbc_list:
        for p in k.split(','):
            p = p.strip()
            if p: kbcs.append(p)
    first_kbc = kbcs[0] if kbcs else ''

    try:
        conn = _ss_conn()
        cur = conn.cursor()
        # Get ALL NVKDs across ALL selected kbcs
        existing = set()
        if kbcs:
            ph = ','.join(['?'] * len(kbcs))
            cur.execute(f'SELECT DISTINCT ma_nvkd FROM kpi_targets WHERE ma_kbc IN ({ph})', kbcs)
            existing = set(r[0] for r in cur.fetchall())

        # Get all NVKDs from system (no ksd filter — include all)
        cur.execute("""SELECT ma_nvkd, ten_nvkd, ma_ql
                       FROM DMNHANVIENKD_VIEW
                       WHERE ma_nvkd IS NOT NULL AND ma_nvkd != ''""")
        all_nv = cur.fetchall()

        bp_map = {}
        cur.execute("""SELECT DISTINCT ma_nvkd, ma_bp FROM DMKHACHHANG_VIEW
                       WHERE ma_nvkd IS NOT NULL AND ma_nvkd != ''
                         AND ma_bp IS NOT NULL AND ma_bp != '' AND ma_bp != 'TN'""")
        for r in cur.fetchall():
            bp_map[r[0]] = r[1]
        conn.close()

        result = []
        for r in all_nv:
            ma = (r[0] or '').strip()
            if not ma or ma in existing:
                continue
            nv_bp = bp_map.get(ma, '')
            if ma_bp and nv_bp != ma_bp:
                continue
            result.append({'ma_nvkd': ma, 'ten_nvkd': (r[1] or '').strip(),
                           'ma_ql': (r[2] or '').strip(), 'ma_bp': nv_bp})
        return jsonify({'ok': True, 'data': result})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@bp.route('/kpi/add-nvkd', methods=['POST'])
@admin_required
def kpi_add_nvkd():
    """Thêm 1 NVKD vào kpi_targets (drag từ unassigned panel)."""
    data = request.get_json(force=True)
    ma_kbc = data.get('ma_kbc', '').strip()
    ma_nvkd = data.get('ma_nvkd', '').strip()
    ten_nvkd = data.get('ten_nvkd', '').strip()
    ma_ql = data.get('ma_ql', '').strip()
    ma_bp = data.get('ma_bp', '').strip()
    if not ma_kbc or not ma_nvkd:
        return jsonify({'ok': False, 'error': 'Thiếu thông tin'}), 400
    nam, thang = _parse_kbc(ma_kbc)
    try:
        conn = _ss_conn()
        cur = conn.cursor()
        cur.execute('SELECT id FROM kpi_targets WHERE ma_kbc = ? AND ma_nvkd = ?', (ma_kbc, ma_nvkd))
        if cur.fetchone():
            conn.close()
            return jsonify({'ok': True, 'message': 'Đã tồn tại'})
        cur.execute('''INSERT INTO kpi_targets
            (ma_bp, ma_kbc, nam, thang, ma_nvkd, ten_nvkd, nguoi_gd, ma_ql)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
            (ma_bp, ma_kbc, nam, thang, ma_nvkd, ten_nvkd, ten_nvkd, ma_ql))
        conn.commit()
        _recalc_paths(conn, ma_kbc)
        conn.close()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@bp.route('/kpi/delete-nvkd', methods=['POST'])
@admin_required
def kpi_delete_nvkd():
    """Xóa 1 hoặc nhiều NVKD khỏi kpi_targets. Con trực tiếp sẽ được gán lên QL cha."""
    data = request.get_json(force=True)
    ma_kbc_list = data.get('ma_kbc', [])
    ma_nvkd = data.get('ma_nvkd', '').strip()
    if isinstance(ma_kbc_list, str):
        ma_kbc_list = [k.strip() for k in ma_kbc_list.split(',') if k.strip()]
    if not ma_nvkd or not ma_kbc_list:
        return jsonify({'ok': False, 'error': 'Thiếu thông tin'}), 400

    try:
        conn = _ss_conn()
        cur = conn.cursor()
        for kbc in ma_kbc_list:
            # Find this NV's parent (ma_ql)
            cur.execute('SELECT ma_ql FROM kpi_targets WHERE ma_kbc = ? AND ma_nvkd = ?', (kbc, ma_nvkd))
            row = cur.fetchone()
            parent_ql = row[0] if row else ''
            # Reassign direct children to parent_ql
            cur.execute('UPDATE kpi_targets SET ma_ql = ?, ldate = GETDATE() WHERE ma_kbc = ? AND ma_ql = ?',
                        (parent_ql or '', kbc, ma_nvkd))
            # Delete the node
            cur.execute('DELETE FROM kpi_targets WHERE ma_kbc = ? AND ma_nvkd = ?', (kbc, ma_nvkd))
            conn.commit()
            _recalc_paths(conn, kbc)
        conn.close()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


# ═══════════════════════════════════════════════
# API: KPI Ratios (tỷ lệ per BP per month)
# ═══════════════════════════════════════════════

@bp.route('/kpi/ratios')
@admin_required
def kpi_ratios_get():
    """Trả về tỷ lệ cho các BP + kỳ."""
    ma_kbc_list = request.args.getlist('ma_kbc')
    ma_bp = request.args.get('ma_bp', '').strip()
    kbcs = []
    for k in ma_kbc_list:
        for p in k.split(','):
            p = p.strip()
            if p: kbcs.append(p)
    if not kbcs:
        return jsonify({'ok': True, 'ratios': {}})
    try:
        conn = _ss_conn()
        cur = conn.cursor()
        ph = ','.join(['?'] * len(kbcs))
        sql = f'SELECT ma_bp, ma_kbc, ratio_dt_cty, ratio_ds_nb, ratio_ds_cty FROM kpi_ratios WHERE ma_kbc IN ({ph})'
        params = list(kbcs)
        if ma_bp:
            sql += ' AND ma_bp = ?'
            params.append(ma_bp)
        cur.execute(sql, params)
        rows = cur.fetchall()
        conn.close()
        # Structure: {VB: {T01-2026: {ratio_dt_cty: 1.04, ...}, T02-2026: ...}}
        ratios = {}
        for r in rows:
            bp_code = r[0] or ''
            kbc = r[1] or ''
            if bp_code not in ratios:
                ratios[bp_code] = {}
            ratios[bp_code][kbc] = {
                'ratio_dt_cty': float(r[2] or 1),
                'ratio_ds_nb': float(r[3] or 1),
                'ratio_ds_cty': float(r[4] or 1),
            }
        return jsonify({'ok': True, 'ratios': ratios})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@bp.route('/kpi/ratios/save', methods=['POST'])
@admin_required
def kpi_ratios_save():
    """Lưu tỷ lệ cho 1 BP + 1 kỳ. Auto-recalc KPI cho toàn bộ NV trong BP+kỳ đó."""
    data = request.get_json(force=True)
    ma_bp = data.get('ma_bp', '').strip()
    ma_kbc = data.get('ma_kbc', '').strip()
    ratio_dt_cty = data.get('ratio_dt_cty', 1)
    ratio_ds_nb = data.get('ratio_ds_nb', 1)
    ratio_ds_cty = data.get('ratio_ds_cty', 1)

    if not ma_bp or not ma_kbc:
        return jsonify({'ok': False, 'error': 'Thiếu ma_bp hoặc ma_kbc'}), 400

    nam, thang = _parse_kbc(ma_kbc)

    try:
        conn = _ss_conn()
        cur = conn.cursor()

        # Upsert ratio
        cur.execute('SELECT id FROM kpi_ratios WHERE ma_bp = ? AND ma_kbc = ?', (ma_bp, ma_kbc))
        if cur.fetchone():
            cur.execute('''UPDATE kpi_ratios SET ratio_dt_cty=?, ratio_ds_nb=?, ratio_ds_cty=?, updated_at=GETDATE()
                           WHERE ma_bp=? AND ma_kbc=?''',
                        (ratio_dt_cty, ratio_ds_nb, ratio_ds_cty, ma_bp, ma_kbc))
        else:
            cur.execute('''INSERT INTO kpi_ratios (ma_bp, ma_kbc, nam, thang, ratio_dt_cty, ratio_ds_nb, ratio_ds_cty)
                           VALUES (?, ?, ?, ?, ?, ?, ?)''',
                        (ma_bp, ma_kbc, nam, thang, ratio_dt_cty, ratio_ds_nb, ratio_ds_cty))

        # Auto-recalc: update kpi_cong_ty, kpi_ds, kpi_ds_cong_ty for all NV in this BP+kbc
        cur.execute('''UPDATE kpi_targets SET
                        kpi_cong_ty = kpi * ?,
                        kpi_ds = kpi * ?,
                        kpi_ds_cong_ty = kpi * ?,
                        ldate = GETDATE()
                       WHERE ma_kbc = ? AND ma_bp = ?''',
                    (ratio_dt_cty, ratio_ds_nb, ratio_ds_cty, ma_kbc, ma_bp))
        updated = cur.rowcount
        conn.commit()
        conn.close()
        return jsonify({'ok': True, 'message': f'Đã lưu tỷ lệ + cập nhật {updated} NV'})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500